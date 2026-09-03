"""
Vendemmia — Análise de Crédito  |  API Backend
FastAPI + BrasilAPI (Receita Federal) + Gemini AI
"""

import asyncio
import asyncpg
import base64
import io
import json
import os
import re
import secrets
import smtplib
import ssl as _ssl_mod
import uuid
import zipfile
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
import openpyxl
import pdfplumber
from dotenv import load_dotenv
from fastapi import Cookie, Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel, model_validator
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

_ENV_FILE = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=_ENV_FILE, override=True)

# ── Rate Limiter ──────────────────────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address, default_limits=["120/minute"])

# ── Diretórios de dados ───────────────────────────────────────────────────────
# No Vercel o sistema de arquivos é read-only; usa /tmp para dados temporários
_IS_VERCEL = bool(os.getenv("VERCEL"))
_TMP_BASE  = Path("/tmp") if _IS_VERCEL else Path(__file__).parent

HISTORICO_DIR = _TMP_BASE / "historico"
HISTORICO_DIR.mkdir(exist_ok=True)

DOCS_DIR = _TMP_BASE / "docs"
DOCS_DIR.mkdir(exist_ok=True)

BACKUPS_DIR = _TMP_BASE / "backups"
BACKUPS_DIR.mkdir(exist_ok=True)

# ── PostgreSQL / Azure ────────────────────────────────────────────────────────
_PG_HOST = os.getenv("PG_HOST", "")
_PG_USER = os.getenv("PG_USER", "")
_PG_PASS = os.getenv("PG_PASS", "")
_PG_DB   = os.getenv("PG_DB",   "vdm_projetos")
_PG_PORT = int(os.getenv("PG_PORT", "5432"))

_PG_POOL: asyncpg.Pool | None = None


def _turso_ok() -> bool:
    """Retorna True se o pool PostgreSQL está disponível."""
    return _PG_POOL is not None


def _mk_ssl() -> _ssl_mod.SSLContext:
    ctx = _ssl_mod.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode    = _ssl_mod.CERT_NONE
    return ctx


async def _pg_init() -> None:
    global _PG_POOL
    if not (_PG_HOST and _PG_USER and _PG_PASS):
        print("[DB] PG_HOST/PG_USER/PG_PASS não configurados — banco desabilitado.")
        return
    try:
        _PG_POOL = await asyncpg.create_pool(
            host=_PG_HOST, port=_PG_PORT,
            user=_PG_USER, password=_PG_PASS,
            database=_PG_DB,
            ssl=_mk_ssl(),
            min_size=1, max_size=8,
            command_timeout=15,
        )
        print(f"[DB] Pool PostgreSQL conectado → {_PG_HOST}/{_PG_DB}")
    except Exception as exc:
        print(f"[DB] Falha ao conectar PostgreSQL: {exc}")


def _sql_pg(sql: str) -> str:
    """Converte placeholders ? para $N (asyncpg)."""
    parts  = sql.split("?")
    result = parts[0]
    for i, part in enumerate(parts[1:], 1):
        result += f"${i}" + part
    return result


async def _turso_query(sql: str, args: list | None = None) -> list[dict]:
    if _PG_POOL is None:
        return []
    async with _PG_POOL.acquire() as conn:
        rows = await conn.fetch(_sql_pg(sql), *(args or []))
        return [dict(row) for row in rows]


async def _turso_exec(sql: str, args: list | None = None) -> None:
    if _PG_POOL is None:
        return
    async with _PG_POOL.acquire() as conn:
        await conn.execute(_sql_pg(sql), *(args or []))


# ── Criação de tabelas + seed de usuários ─────────────────────────────────────

async def _ensure_tables() -> None:
    """Cria tabelas ac_* no PostgreSQL se não existirem e semeia usuários iniciais."""
    if _PG_POOL is None:
        return
    async with _PG_POOL.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_users (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL,
                hashed_password TEXT DEFAULT '',
                role TEXT DEFAULT 'Operações',
                avatar TEXT DEFAULT '',
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT '',
                UNIQUE (email)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_user_passwords (
                email TEXT PRIMARY KEY,
                hashed_password TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_password_reset_tokens (
                token TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used INTEGER DEFAULT 0
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_solicitacoes (
                id TEXT PRIMARY KEY,
                status TEXT DEFAULT 'pendente',
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                data TEXT DEFAULT ''
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_analises (
                id TEXT PRIMARY KEY,
                sol_id TEXT DEFAULT '',
                empresa TEXT DEFAULT '',
                cnpj TEXT DEFAULT '',
                status TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                data TEXT DEFAULT '',
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT ''
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS ac_documents (
                id TEXT PRIMARY KEY,
                sol_id TEXT DEFAULT '',
                tipo TEXT DEFAULT '',
                nome TEXT DEFAULT '',
                content TEXT DEFAULT '',
                mime TEXT DEFAULT '',
                size_bytes INTEGER DEFAULT 0,
                created_at TEXT DEFAULT ''
            )
        """)
        await conn.execute("CREATE INDEX IF NOT EXISTS idx_ac_sol_status   ON ac_solicitacoes(status)")
        await conn.execute("CREATE INDEX IF NOT EXISTS idx_ac_sol_created   ON ac_solicitacoes(created_at DESC)")
        await conn.execute("CREATE INDEX IF NOT EXISTS idx_ac_anal_sol_id   ON ac_analises(sol_id)")
        await conn.execute("CREATE INDEX IF NOT EXISTS idx_ac_anal_created  ON ac_analises(created_at DESC)")
        await conn.execute("CREATE INDEX IF NOT EXISTS idx_ac_docs_sol_id   ON ac_documents(sol_id)")

        # Semeia usuários do users.json se a tabela estiver vazia
        count = await conn.fetchval("SELECT COUNT(*) FROM ac_users")
        if count == 0:
            now_iso = datetime.utcnow().isoformat()
            for u in _load_users():
                await conn.execute(
                    """INSERT INTO ac_users
                       (id, name, email, hashed_password, role, avatar, created_at, updated_at)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                       ON CONFLICT (email) DO NOTHING""",
                    u["id"], u.get("name", ""), u.get("email", ""),
                    u.get("hashed_password", ""), u.get("role", "Operações"),
                    u.get("avatar", ""), now_iso, now_iso,
                )
            print(f"[DB] {len(_load_users())} usuário(s) semeado(s) de users.json")


@asynccontextmanager
async def _lifespan(app):
    await _pg_init()
    await _ensure_tables()
    yield
    if _PG_POOL:
        await _PG_POOL.close()


# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Vendemmia Credit API",
    version="2.0.0",
    docs_url=None,   # desabilita /docs em produção
    redoc_url=None,  # desabilita /redoc em produção
    lifespan=_lifespan,
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ── CORS (origens configuradas via .env) ──────────────────────────────────────
_ORIGINS = [o.strip() for o in os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:8000,http://127.0.0.1:8000"
).split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)


# ── Security Headers Middleware ───────────────────────────────────────────────
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.update({
        "X-Content-Type-Options":  "nosniff",
        "X-Frame-Options":         "DENY",
        "X-XSS-Protection":        "1; mode=block",
        "Referrer-Policy":         "strict-origin-when-cross-origin",
        "Permissions-Policy":      "camera=(), microphone=(), geolocation=()",
        "Cache-Control":           "no-store",
    })
    if os.getenv("HTTPS_ONLY", "false").lower() == "true":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    return response


# ── JWT / Autenticação ────────────────────────────────────────────────────────
_USERS_FILE    = Path(__file__).parent / "users.json"
_pwd_ctx       = CryptContext(schemes=["bcrypt"], deprecated="auto")
_JWT_SECRET    = os.getenv("APP_SECRET_KEY") or secrets.token_hex(32)
_JWT_ALG       = "HS256"
_JWT_HOURS     = int(os.getenv("SESSION_HOURS", "8"))
_SECURE_COOKIE = os.getenv("HTTPS_ONLY", "false").lower() == "true"


def _load_users() -> list:
    # Vercel: usa variável de ambiente USERS_JSON quando o arquivo não existe
    if not _USERS_FILE.exists():
        raw = os.getenv("USERS_JSON", "")
        if raw:
            try:
                return json.loads(raw)
            except Exception:
                return []
        return []
    try:
        return json.loads(_USERS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def _create_token(data: dict) -> str:
    payload = {**data, "exp": datetime.utcnow() + timedelta(hours=_JWT_HOURS)}
    return jwt.encode(payload, _JWT_SECRET, algorithm=_JWT_ALG)


async def _get_current_user(vd_token: Optional[str] = Cookie(default=None)):
    if not vd_token:
        raise HTTPException(status_code=401, detail="Sessão não encontrada — faça login")
    try:
        return jwt.decode(vd_token, _JWT_SECRET, algorithms=[_JWT_ALG])
    except JWTError:
        raise HTTPException(status_code=401, detail="Sessão expirada — faça login novamente")


# ── Role / RLS helpers ───────────────────────────────────────────────────────
# Perfis que podem tomar decisões de crédito (aprovar / negar / encaminhar)
_ROLES_DECISION = {"Financeiro", "Administrador", "Admin", "Diretor"}


def _user_can_decide(user: dict) -> bool:
    role = user.get("role", "")
    return any(r in role for r in _ROLES_DECISION)


def _record_visible_to(record: dict, user: dict) -> bool:
    """RLS: Financeiro/Admin vêem todos; Operações vê apenas seus próprios registros.
    Registros legados sem created_by são visíveis a todos (migração transparente)."""
    if _user_can_decide(user):
        return True
    cb = record.get("created_by") or {}
    if not cb:
        return True
    return cb.get("id") == user.get("sub")


_ROLES_ADMIN = {"Administrador", "Admin"}


async def _require_admin(current_user=Depends(_get_current_user)):
    role = current_user.get("role", "")
    if not any(r in role for r in _ROLES_ADMIN):
        raise HTTPException(403, "Acesso negado — apenas Administradores")
    return current_user


# ── Modelos de autenticação ───────────────────────────────────────────────────
class LoginRequest(BaseModel):
    email: str
    password: str


# ── Endpoints de autenticação ─────────────────────────────────────────────────
@app.post("/api/auth/login")
@limiter.limit("10/minute")
async def auth_login(request: Request, response: Response, body: LoginRequest):
    user = None
    if _turso_ok():
        try:
            rows = await _turso_query(
                "SELECT id, name, email, hashed_password, role, avatar FROM ac_users WHERE email=?",
                [body.email.strip().lower()],
            )
            if rows:
                user = rows[0]
        except Exception:
            pass

    # Fallback to local users.json if Turso is not working/configured or if the user is not found in the DB (for initial migration or fallback)
    if not user:
        users = _load_users()
        user = next((u for u in users if u.get("email", "").lower() == body.email.strip().lower()), None)

    await asyncio.sleep(0.3)  # delay fixo para prevenir timing attacks

    if not user:
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos")

    password_hash = user.get("hashed_password", "")
    # user_passwords tem precedência: é onde o reset de senha salva o novo hash.
    # Verificar sempre, independente de haver hash no registro principal.
    if user.get("id") and _turso_ok():
        try:
            pw_rows = await _turso_query(
                "SELECT hashed_password FROM ac_user_passwords WHERE email=?",
                [body.email.strip().lower()],
            )
            if pw_rows:
                password_hash = pw_rows[0]["hashed_password"]
        except Exception:
            pass

    if not password_hash or not _pwd_ctx.verify(body.password, password_hash):
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos")

    token = _create_token({
        "sub":   user["id"],
        "email": user["email"],
        "name":  user["name"],
        "role":  user.get("role", "Operações"),
    })
    response.set_cookie(
        key="vd_token", value=token,
        httponly=True, samesite="lax",
        secure=_SECURE_COOKIE,
        max_age=_JWT_HOURS * 3600,
        path="/"
    )
    return {
        "ok": True,
        "user": {
            "id":     user["id"],
            "name":   user["name"],
            "email":  user["email"],
            "role":   user.get("role", "Operações"),
            "avatar": user.get("avatar", user["name"][:2].upper()),
        }
    }


@app.post("/api/auth/logout")
async def auth_logout(response: Response):
    response.delete_cookie(key="vd_token", path="/", samesite="lax")
    return {"ok": True}


@app.get("/api/auth/me")
async def auth_me(current_user=Depends(_get_current_user)):
    return {"user": current_user}


# ── Reset de senha ────────────────────────────────────────────────────────────

class ResetRequestModel(BaseModel):
    email: str

class ResetConfirmModel(BaseModel):
    token: str
    password: str


async def _ensure_reset_tables() -> None:
    pass  # tabelas criadas em _ensure_tables() no startup


@app.post("/api/auth/reset-request")
@limiter.limit("5/minute")
async def auth_reset_request(request: Request, body: ResetRequestModel):
    await _ensure_reset_tables()
    email = body.email.strip().lower()

    user = None
    if _turso_ok():
        try:
            rows = await _turso_query(
                "SELECT name, email FROM ac_users WHERE email=?",
                [email],
            )
            if rows:
                user = rows[0]
        except Exception:
            pass

    if not user:
        users = _load_users()
        user  = next((u for u in users if u.get("email", "").lower() == email), None)

    # Responde sempre ok para não revelar se o e-mail existe
    if not user or not _SMTP_HOST:
        return {"ok": True}

    token      = secrets.token_urlsafe(32)
    expires_at = (datetime.utcnow() + timedelta(hours=1)).isoformat()

    await _turso_exec(
        "INSERT INTO ac_password_reset_tokens (token, email, expires_at, used) VALUES (?,?,?,0) ON CONFLICT (token) DO NOTHING",
        [token, email, expires_at],
    )

    base_url  = os.getenv("ALLOWED_ORIGINS", "http://localhost:8000").split(",")[0].strip()
    reset_url = f"{base_url}/login.html?reset={token}"
    nome      = user.get("name", email)

    subject = "Redefinição de senha — Vendemmia Análise de Crédito"
    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f0f0;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f0f0;padding:32px 0;">
  <tr><td align="center">
    <table width="580" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <tr><td style="background:#1e1b4b;padding:22px 32px;">
        <table width="100%" cellpadding="0" cellspacing="0"><tr>
          <td style="color:#fff;font-size:17px;font-weight:700;">Vendemmia</td>
          <td align="right" style="color:rgba(255,255,255,.55);font-size:11px;text-transform:uppercase;letter-spacing:.3px;">Análise de Crédito</td>
        </tr></table>
      </td></tr>
      <tr><td style="padding:32px 32px 24px;">
        <span style="display:inline-block;background:#6366f1;color:#fff;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;padding:4px 11px;border-radius:4px;margin-bottom:20px;">Redefinição de Senha</span>
        <p style="font-size:15px;color:#222;margin:0 0 12px;">Olá, <strong>{nome}</strong>.</p>
        <p style="font-size:14px;color:#555;line-height:1.65;margin:0 0 28px;">
          Recebemos uma solicitação para redefinir a senha da sua conta.<br>
          Clique no botão abaixo para criar uma nova senha. O link é válido por <strong>1 hora</strong>.
        </p>
        <table cellpadding="0" cellspacing="0"><tr><td>
          <a href="{reset_url}" style="display:inline-block;background:linear-gradient(135deg,#422c76,#7c3aed);color:#fff;text-decoration:none;font-size:14px;font-weight:700;padding:14px 32px;border-radius:10px;">
            Redefinir minha senha
          </a>
        </td></tr></table>
        <p style="font-size:12px;color:#999;margin:24px 0 0;line-height:1.7;">
          Se não solicitou, ignore este e-mail — sua senha permanece a mesma.<br>
          Ou copie: <a href="{reset_url}" style="color:#6366f1;word-break:break-all;">{reset_url}</a>
        </p>
      </td></tr>
      <tr><td style="padding:18px 32px;background:#f9f9f9;border-top:1px solid #ebebeb;color:#aaa;font-size:11px;text-align:center;">
        Sistema interno Vendemmia · Não responda este e-mail
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""

    await _send_email(subject, html, [email])
    return {"ok": True}


@app.post("/api/auth/reset-confirm")
@limiter.limit("10/minute")
async def auth_reset_confirm(request: Request, body: ResetConfirmModel):
    await _ensure_reset_tables()

    if not body.token or len(body.token) < 10:
        raise HTTPException(400, "Token inválido")
    if not body.password or len(body.password) < 6:
        raise HTTPException(400, "A senha deve ter no mínimo 6 caracteres")

    rows = await _turso_query(
        "SELECT email, expires_at, used FROM ac_password_reset_tokens WHERE token=?",
        [body.token],
    )
    if not rows:
        raise HTTPException(400, "Link de redefinição inválido ou já utilizado")

    row = rows[0]
    if int(row.get("used") or 0) == 1:
        raise HTTPException(400, "Este link já foi utilizado. Solicite um novo.")

    if datetime.utcnow() > datetime.fromisoformat(row["expires_at"]):
        raise HTTPException(400, "Link expirado. Solicite um novo.")

    email    = row["email"]
    new_hash = _pwd_ctx.hash(body.password)
    now_iso  = datetime.utcnow().isoformat()

    # Atualiza hash nas duas tabelas
    await _turso_exec(
        "UPDATE ac_users SET hashed_password=?, updated_at=? WHERE email=?",
        [new_hash, now_iso, email],
    )
    await _turso_exec(
        "INSERT INTO ac_user_passwords (email, hashed_password, updated_at) VALUES (?,?,?)"
        " ON CONFLICT (email) DO UPDATE SET hashed_password=EXCLUDED.hashed_password, updated_at=EXCLUDED.updated_at",
        [email, new_hash, now_iso],
    )
    await _turso_exec(
        "UPDATE ac_password_reset_tokens SET used=1 WHERE token=?",
        [body.token],
    )
    return {"ok": True}


# ── E-mail ────────────────────────────────────────────────────────────────────
import urllib.parse

_MAILER_DSN = os.getenv("MAILER_DSN", "")
_MAILER_FROM = os.getenv("MAILER_FROM", "")

_SMTP_HOST = ""
_SMTP_PORT = 587
_SMTP_USER = ""
_SMTP_PASS = ""
_SMTP_ENCRYPTION = "tls"

if _MAILER_DSN:
    try:
        _parsed = urllib.parse.urlparse(_MAILER_DSN)
        _q = urllib.parse.parse_qs(_parsed.query)
        
        # Check query parameters first
        _user = _q.get("username", [None])[0]
        _password = _q.get("password", [None])[0]
        _encryption = _q.get("encryption", [None])[0]
        
        # Fallback to standard URL auth
        if not _user and _parsed.username:
            _user = urllib.parse.unquote(_parsed.username)
        if not _password and _parsed.password:
            _password = urllib.parse.unquote(_parsed.password)
            
        _SMTP_HOST = _parsed.hostname or ""
        _SMTP_PORT = _parsed.port or (465 if _parsed.scheme == "smtps" else 587)
        _SMTP_USER = _user or ""
        _SMTP_PASS = _password or ""
        _SMTP_ENCRYPTION = (_encryption or ("ssl" if _parsed.scheme == "smtps" or _SMTP_PORT == 465 else "tls")).lower()
    except Exception:
        pass

# Fallback to individual SMTP_* variables if MAILER_DSN was not set or didn't provide host
if not _SMTP_HOST:
    _SMTP_HOST = os.getenv("SMTP_HOST", "")
    _SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    _SMTP_USER = os.getenv("SMTP_USER", "")
    _SMTP_PASS = os.getenv("SMTP_PASS", "")
    _SMTP_ENCRYPTION = "tls"  # default standard fallback

_FROM_EMAIL = _MAILER_FROM or _SMTP_USER or "noreply@vendemmia.com.br"
_NOTIFY_EMAILS = [e.strip() for e in os.getenv("NOTIFY_EMAILS", "").split(",") if e.strip()]

_STATUS_LABEL = {
    "aprovado":  "Aprovado",
    "negado":    "Negado",
    "em_comite": "Encaminhado ao Comitê",
    "pendente":  "Pendente",
    "em_analise":"Em Análise",
}
_STATUS_COLOR = {
    "aprovado":  "#22c55e",
    "negado":    "#ef4444",
    "em_comite": "#f59e0b",
    "pendente":  "#6366f1",
    "em_analise":"#3b82f6",
}


def _email_html(headline: str, color: str, rows: list) -> str:
    rows_html = "".join(
        f'<tr>'
        f'<td style="padding:7px 0;color:#888;font-size:13px;width:155px;vertical-align:top;">{k}</td>'
        f'<td style="padding:7px 0;color:#111;font-size:13px;font-weight:600;">{v}</td>'
        f'</tr>'
        for k, v in rows if v
    )
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f0f0;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f0f0;padding:32px 0;">
  <tr><td align="center">
    <table width="580" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <tr><td style="background:#1e1b4b;padding:22px 32px;">
        <table width="100%" cellpadding="0" cellspacing="0"><tr>
          <td style="color:#fff;font-size:17px;font-weight:700;letter-spacing:.3px;">Vendemmia</td>
          <td align="right" style="color:rgba(255,255,255,.55);font-size:11px;letter-spacing:.3px;text-transform:uppercase;">Análise de Crédito</td>
        </tr></table>
      </td></tr>
      <tr><td style="padding:28px 32px 4px;">
        <span style="display:inline-block;background:{color};color:#fff;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;padding:4px 11px;border-radius:4px;">{headline}</span>
        <table width="100%" cellpadding="0" cellspacing="0" style="border-top:1px solid #ebebeb;margin-top:18px;">{rows_html}</table>
      </td></tr>
      <tr><td style="padding:20px 32px;background:#f9f9f9;border-top:1px solid #ebebeb;color:#aaa;font-size:11px;text-align:center;">
        Sistema interno Vendemmia &middot; Não responda este e-mail
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""


async def _send_email(subject: str, html: str, to: list, from_name: str = "", from_email: str = "") -> None:
    if not (_SMTP_HOST and to):
        return

    sender_email = from_email or _FROM_EMAIL
    sender_label = f"{from_name} <{sender_email}>" if from_name else sender_email

    def _do():
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = sender_label
        msg["To"]      = ", ".join(to)
        msg.attach(MIMEText(html, "html", "utf-8"))
        
        if _SMTP_ENCRYPTION == "ssl":
            with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT, timeout=20) as s:
                if _SMTP_USER:
                    s.login(_SMTP_USER, _SMTP_PASS)
                s.sendmail(sender_email, to, msg.as_string())
        else:
            with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT, timeout=20) as s:
                s.ehlo()
                if _SMTP_ENCRYPTION == "tls":
                    s.starttls()
                    s.ehlo()
                if _SMTP_USER:
                    s.login(_SMTP_USER, _SMTP_PASS)
                s.sendmail(sender_email, to, msg.as_string())

    try:
        await asyncio.to_thread(_do)
    except Exception:
        pass  # fire-and-forget; não bloqueia o fluxo do usuário


# ── Constantes ────────────────────────────────────────────────────────────────
BRASILAPI  = "https://brasilapi.com.br/api/cnpj/v1"
_SOL_ID_RE = re.compile(r'^[a-zA-Z0-9_\-]{4,64}$')


# ── Modelos ──────────────────────────────────────────────────────────────────

class AnalyzeRequest(BaseModel):
    cnpj: str
    empresa: str
    ramo: Optional[str] = ""
    produto: Optional[str] = ""
    modal: Optional[str] = ""
    segmento: Optional[str] = ""
    origens: Optional[str] = ""
    incoterms: Optional[str] = ""
    tipoOp: Optional[str] = ""
    # Limites
    limiteExportador: Optional[str] = ""
    limiteDesp: Optional[str] = ""
    limiteImp: Optional[str] = ""
    # Volume
    volMes: Optional[str] = ""
    volMesMoeda: Optional[str] = "BRL"
    vol6Meses: Optional[str] = ""
    vol6MesesMoeda: Optional[str] = "USD"
    volPotencial: Optional[str] = ""
    volPotencialMoeda: Optional[str] = "USD"
    # Prazos
    prazoInvoiceData:    Optional[str] = ""
    prazoPrepEmbarque:   Optional[str] = ""
    prazoEmbarque:       Optional[str] = ""
    prazoTransit:        Optional[str] = ""
    prazoDesembaraco:    Optional[str] = ""
    prazoFaturamento:    Optional[str] = ""
    prazoPagtoVendemmia: Optional[str] = ""
    # Fundação
    fundacao: Optional[str] = ""
    # Operação
    importadorFatura: Optional[str] = ""
    consignatario: Optional[str] = ""
    pagtoClienteExp: Optional[str] = ""
    custoFin:     Optional[str] = ""
    custoFinDesc: Optional[str] = ""
    cessao: Optional[str] = ""
    cessaoResp: Optional[str] = ""
    # Financeiro
    rentabilidade: Optional[str] = ""
    rentabilidadeObs: Optional[str] = ""
    custoAdm:       Optional[str] = ""
    custoAdmObs:    Optional[str] = ""
    analyticsValor: Optional[str] = ""
    custoAdmPct:    Optional[str] = ""
    custoAdmBase:   Optional[str] = ""
    custoAdmOutros: Optional[str] = ""
    desconto: Optional[str] = ""
    # Contexto
    comentario: Optional[str] = ""
    sol_id: Optional[str] = ""  # passado para enriquecer o prompt com indicadores contábeis

    @model_validator(mode='before')
    @classmethod
    def _coerce_optional_strings(cls, data: Any) -> Any:
        """Converte arrays/objetos/números para str nos campos opcionais.
        Evita erro 422 quando o frontend envia valores não-string (ex: arrays de seleção múltipla)."""
        if not isinstance(data, dict):
            return data
        required = {'cnpj', 'empresa'}
        for k, v in data.items():
            if k in required:
                continue
            if v is None:
                data[k] = ''
            elif isinstance(v, list):
                data[k] = ', '.join(str(i) for i in v) if v else ''
            elif isinstance(v, dict):
                data[k] = ''
            elif not isinstance(v, str):
                data[k] = str(v)
        return data


# ── Helpers ──────────────────────────────────────────────────────────────────

def clean_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", cnpj)


def br_to_float(val: str) -> float:
    """Converte '250.000,00' → 250000.0"""
    if not val:
        return 0.0
    try:
        return float(val.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def calc_tempo_mercado(fundacao: str) -> str:
    if not fundacao:
        return "Não informado"
    try:
        dt = datetime.strptime(fundacao[:10], "%Y-%m-%d").date()
        anos = (date.today() - dt).days // 365
        meses = ((date.today() - dt).days % 365) // 30
        if anos == 0:
            return f"{meses} {'mês' if meses == 1 else 'meses'}"
        return f"{anos} {'ano' if anos == 1 else 'anos'}" + (f" e {meses} meses" if meses else "")
    except ValueError:
        return fundacao


# ── Consulta Receita Federal (BrasilAPI + fallback CNPJA) ────────────────────

_CNPJA_URL = "https://open.cnpja.com/office"

def _normalize_cnpja(d: dict) -> dict:
    """Mapeia resposta da CNPJA para o formato BrasilAPI."""
    addr = d.get("address") or {}
    company = d.get("company") or {}
    status  = d.get("status")  or {}
    nature  = company.get("nature") or {}
    act     = (d.get("mainActivity") or {})
    return {
        "razao_social":                    company.get("name", ""),
        "nome_fantasia":                   d.get("alias") or "",
        "descricao_situacao_cadastral":    status.get("text", ""),
        "data_situacao_cadastral":         d.get("statusDate", ""),
        "data_abertura":                   d.get("founded", ""),
        "cnae_fiscal_descricao":           act.get("text", ""),
        "natureza_juridica":               nature.get("text", ""),
        "logradouro":                      addr.get("street", ""),
        "numero":                          str(addr.get("number", "")),
        "complemento":                     addr.get("details", ""),
        "bairro":                          addr.get("district", ""),
        "municipio":                       addr.get("city", ""),
        "uf":                              addr.get("state", ""),
        "cep":                             str(addr.get("zip", "")),
        "email":                           d.get("emails", [{}])[0].get("address", "") if d.get("emails") else "",
        "telefone":                        d.get("phones", [{}])[0].get("number", "")  if d.get("phones")  else "",
        "_fonte":                          "cnpja",
    }


async def fetch_receita(cnpj: str) -> dict:
    clean = clean_cnpj(cnpj)
    if len(clean) != 14:
        return {"status": "invalid", "data": {}, "error": "CNPJ deve ter 14 dígitos"}

    # 1ª tentativa — BrasilAPI (com retry em caso de 429)
    for attempt in range(2):
        try:
            async with httpx.AsyncClient(timeout=12.0) as c:
                resp = await c.get(f"{BRASILAPI}/{clean}")
            if resp.status_code == 200:
                return {"status": "ok", "data": resp.json()}
            if resp.status_code == 404:
                return {"status": "not_found", "data": {}, "error": "CNPJ não encontrado na Receita Federal"}
            if resp.status_code == 429 and attempt == 0:
                await asyncio.sleep(2)
                continue
            # outro erro de status — tenta fallback
            break
        except httpx.TimeoutException:
            break
        except Exception:
            break

    # Fallback — CNPJA (API pública, sem chave)
    try:
        async with httpx.AsyncClient(timeout=15.0) as c:
            resp = await c.get(
                f"{_CNPJA_URL}/{clean}",
                headers={"User-Agent": "vendemmia-analise-credito/2.0"},
            )
        if resp.status_code == 200:
            return {"status": "ok", "data": _normalize_cnpja(resp.json())}
        if resp.status_code == 404:
            return {"status": "not_found", "data": {}, "error": "CNPJ não encontrado na Receita Federal"}
    except httpx.TimeoutException:
        return {"status": "timeout", "data": {}, "error": "Timeout ao consultar Receita Federal"}
    except Exception as exc:
        return {"status": "error", "data": {}, "error": str(exc)}

    return {"status": "error", "data": {}, "error": "Serviço de consulta temporariamente indisponível. Tente novamente em alguns minutos."}


# ── Prompt ───────────────────────────────────────────────────────────────────

def build_prompt(req: AnalyzeRequest, receita: dict, contabil_result: Optional[dict] = None) -> str:
    d = receita.get("data", {})
    bureau_ok = receita.get("status") == "ok"

    # Quadro societário
    qsa_lines = ""
    if d.get("qsa"):
        rows = []
        for s in d["qsa"]:
            rows.append(
                f"  • {s.get('nome_socio','?')} | "
                f"{s.get('percentual_capital_social','?')}% | "
                f"Faixa etária: {s.get('faixa_etaria','?')} | "
                f"Entrada: {s.get('data_entrada_sociedade','?')}"
            )
        qsa_lines = "\n".join(rows)

    # Seção Receita Federal
    if bureau_ok and d:
        abertura = d.get("data_inicio_atividade", "")
        try:
            dt = datetime.strptime(abertura, "%Y-%m-%d").date()
            anos_rf = (date.today() - dt).days // 365
            tempo_rf = f"{anos_rf} anos"
        except Exception:
            tempo_rf = abertura

        receita_section = f"""
## DADOS DA RECEITA FEDERAL (BrasilAPI — tempo real)
- Razão Social: {d.get('razao_social', '—')}
- Nome Fantasia: {d.get('nome_fantasia') or '—'}
- Situação Cadastral: {d.get('descricao_situacao_cadastral', '—')}
- Data da Situação: {d.get('data_situacao_cadastral', '—')}
- Abertura: {abertura} ({tempo_rf})
- Capital Social: R$ {d.get('capital_social', 0):,.2f}
- Natureza Jurídica: {d.get('descricao_natureza_juridica', '—')}
- Porte: {d.get('descricao_porte', '—')}
- CNAE Principal: {d.get('cnae_fiscal', '—')} — {d.get('cnae_fiscal_descricao', '—')}
- Simples Nacional: {'Sim' if d.get('opcao_pelo_simples') else 'Não'}
- MEI: {'Sim' if d.get('opcao_pelo_mei') else 'Não'}
- UF / Município: {d.get('uf', '—')} / {d.get('municipio', '—')}

## QUADRO SOCIETÁRIO (QSA)
{qsa_lines or 'Não disponível'}
"""
    else:
        receita_section = f"""
## DADOS DA RECEITA FEDERAL
Status: {receita.get('status')} — {receita.get('error', 'Indisponível')}
CNPJ informado: {req.cnpj}
"""

    # Cálculo de exposição total
    exp_total = br_to_float(req.limiteExportador) + br_to_float(req.limiteDesp) + br_to_float(req.limiteImp)
    exp_str = f"R$ {exp_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    tempo_mercado  = calc_tempo_mercado(req.fundacao) if req.fundacao else "Não informado"
    obs_analista   = f"## OBSERVAÇÕES DO ANALISTA\n{req.comentario}" if req.comentario else ""
    contabil_bloco = _build_contabil_section(contabil_result) if contabil_result and contabil_result.get("periodo2") else ""

    return f"""Você é um analista de crédito sênior especializado em empresas importadoras no Brasil,
trabalhando na Vendemmia — empresa de logística de importação (Trading/Account).

O crédito que a Vendemmia concede representa o risco de pagar adiantado fretes internacionais,
despesas alfandegárias e impostos de importação, sendo reembolsado pelo cliente posteriormente.
Inadimplência = Vendemmia absorve o custo integralmente.

{receita_section}

## DADOS DA SOLICITAÇÃO (informados pelo time de operações)
- Empresa: {req.empresa}
- CNPJ: {req.cnpj}
- Ramo de atividade: {req.ramo or 'Não informado'}
- Produto importado: {req.produto or 'Não informado'}
- Modal logístico: {req.modal or 'Não informado'}
- Segmento: {req.segmento or 'Não informado'}
- Principais origens: {req.origens or 'Não informado'}
- Incoterms: {req.incoterms or 'Não informado'}
- Tipo de operação: {req.tipoOp or 'Não informado'}
- Tempo de mercado (declarado): {tempo_mercado}

## LIMITES SOLICITADOS
- Exportador (câmbio/mercadoria): R$ {req.limiteExportador or '0'}
- Despesas alfandegárias: R$ {req.limiteDesp or '0'}
- Impostos de importação: R$ {req.limiteImp or '0'}
- EXPOSIÇÃO TOTAL: {exp_str}

## VOLUME DE NEGÓCIOS
- Volume mensal estimado: {req.volMes or '—'} {req.volMesMoeda or ''}
- Volume 6 meses estimado: {req.vol6Meses or '—'} {req.vol6MesesMoeda or ''}
- Potencial anual: {req.volPotencial or '—'} {req.volPotencialMoeda or ''}
- Rentabilidade Vendemmia: {req.rentabilidade or '—'}%{(' (' + req.rentabilidadeObs + ')') if req.rentabilidadeObs else ''}
- Plataforma Analytics: R$ {req.analyticsValor or '500,00'} por processo
- Taxa administrativa: {(req.custoAdmPct + '% sobre ' + (req.custoAdmBase or 'CIF') + ((' — ' + req.custoAdmObs) if req.custoAdmObs else '')) if req.custoAdmPct else '—'}
{('- Outros custos adm.: ' + req.custoAdmOutros) if req.custoAdmOutros else ''}
- Desconto sobre tabela: {req.desconto or '0'}%

## ESTRUTURA DA OPERAÇÃO
- Importador da fatura: {req.importadorFatura or '—'}
- Consignatário: {req.consignatario or '—'}
- Pagamento ao exportador: {req.pagtoClienteExp or '—'}
- Custo financeiro cobrado do cliente: {req.custoFin or '—'}{(' (' + req.custoFinDesc + ')') if req.custoFinDesc else ''}
- Cessão de crédito: {req.cessao or '—'}{(' — Responsável: ' + req.cessaoResp) if req.cessaoResp else ''}

## PRAZOS
- Invoice: {req.prazoInvoiceData or '—'}
- Preparação para embarque: {req.prazoPrepEmbarque or '—'} dias
- Embarque: {req.prazoEmbarque or '—'} dias
- Trânsito internacional: {req.prazoTransit or '—'} dias
- Desembaraço aduaneiro: {req.prazoDesembaraco or '—'} dias
- Faturamento: {req.prazoFaturamento or '—'} dias
- Pagamento à Vendemmia: {req.prazoPagtoVendemmia or '—'} dias
{contabil_bloco}
Retorne APENAS um JSON válido, sem texto adicional antes ou depois:

{{
  "score": <inteiro 0-100; 100 = risco mínimo / empresa excelente>,
  "classificacao": "<AAA|AA|A|BB|B|CC|C|D>",
  "recomendacao": "<aprovar|negar|revisar>",
  "limite_recomendado_exportador": "<R$ formatado ou 'Não recomendado'>",
  "limite_recomendado_desp": "<R$ formatado ou 'Não recomendado'>",
  "limite_recomendado_imp": "<R$ formatado ou 'Não recomendado'>",
  "exposicao_total_recomendada": "<R$ formatado>",
  "prazo_recomendado": <30|45|60|90|120>,
  "resumo_executivo": "<2-3 frases objetivas e diretas>",
  "pontos_positivos": ["<ponto 1>", "<ponto 2>"],
  "pontos_atencao": ["<ponto 1>", "<ponto 2>"],
  "alertas_criticos": [],
  "analise_cadastral": "<análise da situação na Receita Federal em 2-3 frases>",
  "analise_societaria": "<análise do quadro societário, perfil dos sócios, concentração de capital>",
  "analise_proporcionalidade": "<análise da proporcionalidade entre exposição total, capital social e volume declarado>",
  "analise_operacional": "<análise dos riscos operacionais: modal, origens, prazo de trânsito, tipo de operação>",
  "fundamentacao": "<análise completa em 3-5 parágrafos cobrindo: (1) situação cadastral e societária, (2) proporcionalidade do crédito, (3) riscos operacionais de importação, (4) recomendação final com condições>"
}}

Diretrizes de pontuação (orientativas):
- Situação ATIVA na Receita Federal: +25 pts
- Empresa > 5 anos: +20 pts | 2-5 anos: +10 pts | < 2 anos: -10 pts
- Capital social ≥ exposição total: +15 pts | ≥ 50%: +8 pts | < 20%: -10 pts
- CNAE compatível com produto importado: +10 pts
- Simples Nacional: -3 pts | MEI: -25 pts (limitar a R$ 10.000)
- Situação INAPTA ou BAIXADA: score ≤ 15, recomendação obrigatoriamente "negar"
- Sócio único + empresa < 1 ano: alerta crítico
- Exposição > 3× volume mensal declarado: alerta crítico
{obs_analista}
"""


# ── Endpoints ────────────────────────────────────────────────────────────────

def _extract_json(text: str) -> Optional[dict]:
    """Extrai o objeto JSON da resposta do Claude."""
    # Estratégia 1: parse direto
    try:
        obj = json.loads(text.strip())
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass

    # Estratégia 2: segmentos entre triple-backticks
    parts = text.split("```")
    for part in parts:
        candidate = part.strip()
        if candidate.startswith("json"):
            candidate = candidate[4:].strip()
        if candidate.startswith("{"):
            try:
                obj = json.loads(candidate)
                if isinstance(obj, dict):
                    return obj
            except Exception:
                pass

    # Estratégia 3: primeiro { ao último }
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end > start:
        try:
            obj = json.loads(text[start:end + 1])
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass

    return None


def _load_key() -> str:
    """Carrega e limpa a chave Gemini do .env."""
    load_dotenv(dotenv_path=_ENV_FILE, override=True)
    raw = os.environ.get("GEMINI_API_KEY", "").strip()
    # Remove aspas acidentais que python-dotenv pode manter em edge cases
    raw = raw.strip('"').strip("'").strip()
    return raw


def _load_gemini_model() -> str:
    """Retorna o nome do modelo Gemini — configurável via GEMINI_MODEL no .env."""
    load_dotenv(dotenv_path=_ENV_FILE, override=True)
    return os.environ.get("GEMINI_MODEL", "gemini-1.5-flash").strip().strip('"').strip("'") or "gemini-1.5-flash"


def _load_anthropic_key() -> str:
    """Carrega e limpa a chave Anthropic do .env."""
    load_dotenv(dotenv_path=_ENV_FILE, override=True)
    raw = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    return raw.strip('"').strip("'").strip()


# ── Extração de BP/DRE por regex (sem IA) ────────────────────────────────────

_BR_NUM_RE = re.compile(
    r'\(\s*[\d]+(?:\.[\d]{3})*(?:,\d{1,2})?\s*\)'  # (1.234,56) negativo
    r'|[\d]{1,3}(?:\.[\d]{3})+(?:,\d{1,2})?'       # 1.234.567,89
    r'|[\d]+,\d{2}'                                  # 1234,56
)

def _parse_br_num(s: str) -> Optional[float]:
    """Converte string de número brasileiro para float."""
    if not s:
        return None
    s = s.strip()
    negative = s.startswith('(') and s.endswith(')')
    s = s.strip('()')
    s = s.replace('.', '').replace(',', '.')
    try:
        v = float(s)
        return (-v if negative else v) if v != 0 else None
    except ValueError:
        return None

def _nums_from_line(text: str) -> list:
    """Extrai lista de floats de um trecho de texto (formato brasileiro)."""
    return [v for v in (_parse_br_num(m) for m in _BR_NUM_RE.findall(text)) if v is not None]

_BP_PATS: dict = {
    'disponibilidade':       [r'caixa\s+e\s+equiv', r'disponibilidade', r'caixa\s+e\s+banco'],
    'contas_receber':        [r'contas\s+a\s+receber', r'clientes\s*$', r'duplicatas\s+a\s+receber'],
    'estoques':              [r'\bestoques?\b'],
    'impostos_recuperar':    [r'impostos?\s+a\s+recuperar', r'tributos?\s+a\s+recuperar', r'ativo\s+fiscal\b'],
    'outros_ac':             [r'outros\s+ativos?\s+circulant', r'outros\s+cr[eé]ditos\s+circulant'],
    'outros_creditos':       [r'outros\s+cr[eé]ditos', r'realizável\s+a\s+longo'],
    'imobilizado':           [r'\bimobilizado\b'],
    'investimentos':         [r'\binvestimentos?\b'],
    'outros_anc':            [r'intang[ií]vel', r'outros\s+ativos?\s+n[ãa]o\s+circulant'],
    'fornecedores':          [r'\bfornecedores?\b'],
    'adiantamento_clientes': [r'adiantamento.*cliente', r'receitas?\s+diferidas?'],
    'impostos_pagar':        [r'impostos?\s+(a\s+)?pagar', r'tributos?\s+(a\s+)?pagar', r'obriga[çc][õo]es?\s+fiscais?'],
    'emprestimos_cp':        [r'empr[eé]stimos?.*(?:curto|circulant)', r'financiamentos?\s+circulant'],
    'outros_pc':             [r'outros\s+passivos?\s+circulant'],
    'emprestimos_lp':        [r'empr[eé]stimos?.*(?:longo|n[ãa]o\s+circulant)', r'financiamentos?\s+n[ãa]o\s+circulant'],
    'outros_pnc':            [r'outros\s+passivos?\s+n[ãa]o\s+circulant'],
    'patrimonio_liquido':    [r'patrim[ôo]nio\s+l[íi]quido\s*$', r'total.*patrim[ôo]nio'],
}

_DRE_PATS: dict = {
    'receita_bruta':         [r'receita\s+(?:operacional\s+)?bruta', r'receita\s+de\s+vendas'],
    'deducoes':              [r'dedu[çc][õo]es?', r'impostos?\s+sobre\s+(?:venda|fatura|receita)'],
    'receita_liquida':       [r'receita\s+(?:operacional\s+)?l[íi]quida', r'vendas?\s+l[íi]quidas?'],
    'cpv':                   [r'custo\s+(?:dos?\s+)?(?:produtos?|mercadoria|servi[çc]os?)\s+vendidos?', r'\bcpv\b', r'\bcmo\b'],
    'lucro_bruto':           [r'lucro\s+bruto'],
    'despesas_operacionais': [r'despesas?\s+operacionais?', r'despesas?\s+(?:com\s+)?vendas?', r'despesas?\s+(?:gerais?|administrativas?)'],
    'ebitda':                [r'\bebitda\b', r'\blajida\b'],
    'resultado_financeiro':  [r'resultado\s+financeiro', r'receitas?\s+financeiras?'],
    'lucro_antes_ir':        [r'lucro\s+antes\s+(?:do\s+)?imposto', r'\blair\b'],
    'ir_csll':               [r'imposto\s+de\s+renda', r'ir\s+e\s+csll', r'irpj'],
    'lucro_liquido':         [r'lucro\s+(?:l[íi]quido|do\s+per[íi]odo)', r'resultado\s+(?:l[íi]quido|do\s+per[íi]odo)'],
}

_PERIOD_LABEL_RE = re.compile(
    r'(?:dez|jun|mar|set|jan|fev|abr|mai|jul|ago|out|nov)[./\s]\d{4}'
    r'|\d{2}/\d{2}/\d{4}',
    re.IGNORECASE,
)


def _extract_contabil_regex(texto: str) -> Optional[dict]:
    """
    Extrai BP e DRE do texto pdfplumber usando regex — sem IA.
    Suporta: (a) tabelas com pipe (duas colunas = dois períodos)
             (b) documentos separados por '=== DOCUMENTO: ==='
             (c) texto simples com um valor por linha
    """
    # Separar seções de documento
    parts = re.split(r'(=== DOCUMENTO:.*?===)', texto)
    docs: list[tuple[str, str]] = []  # (nome, texto)
    cur_name = ''
    for p in parts:
        m = re.match(r'=== DOCUMENTO:\s*(.+?)\s*===', p)
        if m:
            cur_name = m.group(1)
        elif p.strip():
            docs.append((cur_name, p.strip()))

    if not docs:
        docs = [('', texto)]

    def _period_label(name: str, text: str) -> str:
        dates = _PERIOD_LABEL_RE.findall(text[:800])
        if dates:
            return dates[0]
        y = re.search(r'20\d{2}', name)
        return y.group() if y else ''

    def _parse_section(text: str) -> dict:
        """Retorna bp1, dre1, bp2, dre2 e has_p2 (dois períodos na mesma tabela)."""
        bp1  = {k: None for k in _BP_PATS}
        bp2  = {k: None for k in _BP_PATS}
        dre1 = {k: None for k in _DRE_PATS}
        dre2 = {k: None for k in _DRE_PATS}
        has_p2 = False

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            ll = line.lower()

            if '|' in line:
                cols = [c.strip() for c in line.split('|')]
                label = cols[0].lower()
                vals: list = []
                for c in cols[1:]:
                    vals.extend(_nums_from_line(c))
                v1 = vals[0] if len(vals) >= 1 else None
                v2 = vals[1] if len(vals) >= 2 else None
                if v2 is not None:
                    has_p2 = True
            else:
                label = ll
                nums = _nums_from_line(line)
                if not nums:
                    continue
                v1, v2 = nums[-1], None

            for field, pats in _BP_PATS.items():
                if bp1[field] is None and any(re.search(p, label) for p in pats):
                    bp1[field] = v1
                    if v2 is not None:
                        bp2[field] = v2
                    break
            else:
                for field, pats in _DRE_PATS.items():
                    if dre1[field] is None and any(re.search(p, label) for p in pats):
                        dre1[field] = v1
                        if v2 is not None:
                            dre2[field] = v2
                        break

        return dict(bp1=bp1, bp2=bp2, dre1=dre1, dre2=dre2, has_p2=has_p2)

    r0 = _parse_section(docs[0][1])

    if r0['has_p2']:
        # Dois períodos na mesma tabela
        text0 = docs[0][1]
        dates = _PERIOD_LABEL_RE.findall(text0[:800])
        lbl1 = dates[0] if len(dates) > 0 else _period_label(docs[0][0], text0)
        lbl2 = dates[1] if len(dates) > 1 else 'Período 2'
        bp1, dre1 = r0['bp1'], r0['dre1']
        bp2, dre2 = r0['bp2'], r0['dre2']
    elif len(docs) >= 2:
        # Documentos separados, um por período
        r1 = _parse_section(docs[1][1])
        bp1, dre1 = r0['bp1'], r0['dre1']
        bp2, dre2 = r1['bp1'], r1['dre1']
        lbl1 = _period_label(docs[0][0], docs[0][1]) or 'Período 1'
        lbl2 = _period_label(docs[1][0], docs[1][1]) or 'Período 2'
    else:
        bp1, dre1 = r0['bp1'], r0['dre1']
        bp2 = {k: None for k in _BP_PATS}
        dre2 = {k: None for k in _DRE_PATS}
        lbl1 = _period_label(docs[0][0], docs[0][1]) or 'Período 1'
        lbl2 = ''

    useful = [v for v in list(bp1.values()) + list(dre1.values()) if v is not None]
    if len(useful) < 3:
        return None

    return {
        'periodo1_label':    lbl1,
        'periodo2_label':    lbl2,
        'bp1':               bp1,
        'bp2':               bp2,
        'dre1':              dre1,
        'dre2':              dre2,
        'fator_risco':       0.10,
        'fator_multiplicador': 1.1,
    }


def _claude_extract_json(texto: str, prompt_schema: str, anthropic_key: str) -> Optional[dict]:
    """
    Usa a API do Claude (Anthropic) para extrair dados estruturados de texto financeiro.
    Retorna dict com os dados ou None em caso de falha.
    """
    prompt = f"""{prompt_schema}

DOCUMENTOS:
{texto[:20000]}

Retorne APENAS o JSON válido, sem markdown, sem texto extra."""

    try:
        resp = httpx.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": anthropic_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 4096,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=60.0,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"Anthropic HTTP {resp.status_code}: {resp.text[:200]}")
        data = resp.json()
        raw_text = data["content"][0]["text"]
        return _extract_json(raw_text)
    except Exception as exc:
        raise RuntimeError(f"Anthropic API error: {exc}") from exc


_GEMINI_FALLBACK_MODELS = [
    # Tentados em ordem se o modelo configurado falhar por NOT_FOUND
    "gemini-2.0-flash-exp",
    "gemini-2.5-flash-preview-05-20",
    "gemini-2.5-flash-lite-preview-06-17",
    "gemini-1.5-flash",
    "gemini-1.5-pro",
    "gemini-1.0-pro",
]


def _gemini_rest_call(key: str, model: str, prompt: str) -> str:
    """Chama a Gemini REST API diretamente, tentando v1beta e depois v1."""
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 8192},
    }
    params = {"key": key}

    for api_version in ("v1beta", "v1"):
        url = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
        resp = httpx.post(url, params=params, json=body, timeout=120)
        if resp.status_code == 200:
            data = resp.json()
            return data["candidates"][0]["content"]["parts"][0]["text"]

        try:
            err = resp.json().get("error", {})
        except Exception:
            err = {}
        code = err.get("code", resp.status_code)
        msg  = err.get("message", resp.text)
        status_str = err.get("status", "")

        # Erros definitivos — não adianta trocar versão de API nem modelo
        if code in (400, 403) or status_str in ("INVALID_ARGUMENT", "PERMISSION_DENIED"):
            raise RuntimeError(f"[ClientError] {code} {status_str}: {msg}")
        if code == 429 or status_str == "RESOURCE_EXHAUSTED":
            raise RuntimeError(f"[ClientError] 429 RESOURCE_EXHAUSTED: {msg}")

        # 404 NOT_FOUND → tenta próxima versão de API
        if code == 404:
            continue

        raise RuntimeError(f"[ClientError] {code} {status_str}: {msg}")

    raise RuntimeError(f"[ClientError] Modelo '{model}' não encontrado em v1beta nem em v1")


def _gemini_generate(key: str, prompt: str) -> str:
    """Chama Gemini com fallback automático de modelo se NOT_FOUND."""
    configured = _load_gemini_model()

    models_to_try = [configured]
    for m in _GEMINI_FALLBACK_MODELS:
        if m != configured:
            models_to_try.append(m)

    last_err: Exception = RuntimeError("Gemini indisponível")
    for model in models_to_try:
        try:
            return _gemini_rest_call(key, model, prompt)
        except RuntimeError as exc:
            msg = str(exc)
            last_err = exc
            # Erros de key/quota: não adianta tentar outros modelos
            if any(x in msg for x in ("INVALID_ARGUMENT", "PERMISSION_DENIED", "RESOURCE_EXHAUSTED", "400", "403", "429")):
                break
            # NOT_FOUND → tenta próximo modelo na lista
            continue

    raise last_err


# ── Modelo histórico ─────────────────────────────────────────────────────────

class HistoricoSaveRequest(BaseModel):
    solicitacao_id: str
    empresa: str
    cnpj: str
    status_solicitacao: Optional[str] = ""
    solicitante: Optional[str] = ""
    data_solicitacao: Optional[str] = ""
    dados_solicitacao: Optional[Dict[str, Any]] = None
    receita_federal: Optional[Dict[str, Any]] = None
    analise_ia: Optional[Dict[str, Any]] = None
    modelo_ia: Optional[str] = "claude-sonnet-4-6"
    # Timestamps das etapas do processo
    solicitacao_criada_at: Optional[str] = None
    rf_consultada_at: Optional[str] = None
    analise_ia_at: Optional[str] = None


def _hist_id_safe(hist_id: str) -> bool:
    return bool(re.match(r'^[a-f0-9\-]{36}$', hist_id))


@app.post("/api/historico")
async def salvar_historico(entry: HistoricoSaveRequest, current_user=Depends(_get_current_user)):
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    now_iso = datetime.now().isoformat()
    hist_id = str(uuid.uuid4())
    created_by = {
        "id":    current_user.get("sub", ""),
        "email": current_user.get("email", ""),
        "name":  current_user.get("name", ""),
    }
    record = {
        "id": hist_id,
        "solicitacao_id": entry.solicitacao_id,
        "empresa": entry.empresa,
        "cnpj": entry.cnpj,
        "status_solicitacao": entry.status_solicitacao or "",
        "solicitante": current_user.get("name", entry.solicitante or ""),
        "data_solicitacao": entry.data_solicitacao or "",
        "dados_solicitacao": entry.dados_solicitacao or {},
        "receita_federal": entry.receita_federal or {},
        "analise_ia": entry.analise_ia or {},
        "modelo_ia": entry.modelo_ia or "claude-sonnet-4-6",
        "decisao_analista": None,
        "created_by": created_by,
        "timestamps": {
            "solicitacao_criada_at": entry.solicitacao_criada_at or entry.data_solicitacao or "",
            "rf_consultada_at":      entry.rf_consultada_at or now_iso,
            "analise_ia_at":         entry.analise_ia_at    or now_iso,
            "historico_salvo_at":    now_iso,
            "decisao_at":            None,
        },
    }
    await _turso_exec(
        "INSERT INTO ac_analises (id, sol_id, empresa, cnpj, status, created_by, data, created_at, updated_at) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        [hist_id, entry.solicitacao_id or "", entry.empresa, entry.cnpj,
         record["status_solicitacao"] or "pendente",
         json.dumps(created_by, ensure_ascii=False),
         json.dumps(record, ensure_ascii=False),
         now_iso, now_iso],
    )
    return {"id": hist_id, "salvo_em": now_iso}


@app.get("/api/historico")
async def listar_historico(
    cnpj: Optional[str] = None,
    empresa: Optional[str] = None,
    limit: int = 200,
    current_user=Depends(_get_current_user),
):
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    rows = await _turso_query(
        "SELECT data FROM ac_analises ORDER BY created_at DESC LIMIT ?", [limit * 5]
    )
    entries: List[dict] = []
    for row in rows:
        if len(entries) >= limit:
            break
        try:
            data = json.loads(row["data"])
            if not _record_visible_to(data, current_user):
                continue
            if cnpj:
                if re.sub(r"\D", "", cnpj) not in re.sub(r"\D", "", data.get("cnpj", "")):
                    continue
            if empresa:
                if empresa.lower() not in data.get("empresa", "").lower():
                    continue
            ai  = data.get("analise_ia") or {}
            rf  = (data.get("receita_federal") or {}).get("data") or {}
            ts  = data.get("timestamps") or {}
            dec = data.get("decisao_analista") or {}
            entries.append({
                "id": data["id"],
                "cnpj": data["cnpj"],
                "empresa": data["empresa"],
                "score": ai.get("score"),
                "classificacao": ai.get("classificacao"),
                "recomendacao": ai.get("recomendacao"),
                "resumo_executivo": (ai.get("resumo_executivo") or "")[:250],
                "alertas_criticos": len(ai.get("alertas_criticos") or []),
                "pontos_positivos": len(ai.get("pontos_positivos") or []),
                "limite_recomendado_exportador": ai.get("limite_recomendado_exportador"),
                "exposicao_total_recomendada": ai.get("exposicao_total_recomendada"),
                "prazo_recomendado": ai.get("prazo_recomendado"),
                "rf_situacao": rf.get("descricao_situacao_cadastral"),
                "rf_abertura": rf.get("data_inicio_atividade"),
                "solicitante": data.get("solicitante"),
                "data_solicitacao": data.get("data_solicitacao"),
                "status_solicitacao": data.get("status_solicitacao"),
                "decisao_analista": dec,
                "modelo_ia": data.get("modelo_ia"),
                "solicitacao_id": data.get("solicitacao_id"),
                "receita_federal": data.get("receita_federal") or {},
                "analise_ia":      data.get("analise_ia")      or {},
                "dados_solicitacao": data.get("dados_solicitacao") or {},
                "timestamps": {
                    "solicitacao_criada_at": ts.get("solicitacao_criada_at") or data.get("data_solicitacao"),
                    "rf_consultada_at":      ts.get("rf_consultada_at"),
                    "analise_ia_at":         ts.get("analise_ia_at"),
                    "historico_salvo_at":    ts.get("historico_salvo_at"),
                    "decisao_at":            ts.get("decisao_at") or dec.get("decisao_at"),
                },
            })
        except Exception:
            continue
    return {"total": len(entries), "entries": entries}


@app.get("/api/historico/{hist_id}")
async def buscar_historico(hist_id: str, current_user=Depends(_get_current_user)):
    if not _hist_id_safe(hist_id):
        raise HTTPException(400, "ID inválido")
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    rows = await _turso_query("SELECT data FROM ac_analises WHERE id=?", [hist_id])
    if not rows:
        raise HTTPException(404, "Análise não encontrada")
    data = json.loads(rows[0]["data"])
    if not _record_visible_to(data, current_user):
        raise HTTPException(404, "Análise não encontrada")
    return data


@app.patch("/api/historico/{hist_id}/decisao")
async def atualizar_decisao(hist_id: str, body: Dict[str, Any], current_user=Depends(_get_current_user)):
    if not _user_can_decide(current_user):
        raise HTTPException(403, "Acesso negado — apenas Financeiro e Administrador podem registrar decisões de crédito")
    if not _hist_id_safe(hist_id):
        raise HTTPException(400, "ID inválido")
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    rows = await _turso_query("SELECT data FROM ac_analises WHERE id=?", [hist_id])
    if not rows:
        raise HTTPException(404, "Análise não encontrada")
    data    = json.loads(rows[0]["data"])
    now_iso = datetime.now().isoformat()
    decisao_at = body.get("decisao_at") or now_iso
    decisao_payload = {
        "status":          body.get("status", ""),
        "limiteAprovado":  body.get("limiteAprovado", ""),
        "limiteDesp":      body.get("limiteDesp", ""),
        "limiteImp":       body.get("limiteImp", ""),
        "prazoAprovado":   body.get("prazoAprovado", ""),
        "analistaObs":     body.get("analistaObs", ""),
        "parecerTecnico":  body.get("parecerTecnico", ""),
        "decisaoAnalista": current_user.get("name", body.get("decisaoAnalista", "Analista")),
        "decisao_at":      decisao_at,
    }
    data["decisao_analista"] = decisao_payload
    if decisao_payload["status"]:
        data["status_solicitacao"] = decisao_payload["status"]
    ds = data.get("dados_solicitacao") or {}
    ds.update({k: decisao_payload[k] for k in decisao_payload})
    data["dados_solicitacao"] = ds
    if "timestamps" not in data:
        data["timestamps"] = {}
    data["timestamps"]["decisao_at"] = decisao_at
    data["atualizado_em"] = now_iso
    await _turso_exec(
        "UPDATE ac_analises SET data=?, status=?, updated_at=? WHERE id=?",
        [json.dumps(data, ensure_ascii=False), decisao_payload["status"] or data.get("status", "pendente"), now_iso, hist_id],
    )

    # Disparo automático de e-mail ao solicitante quando a análise é finalizada
    sol_id = data.get("solicitacao_id")
    if sol_id and _SMTP_HOST:
        try:
            sol_rows = await _turso_query(
                "SELECT created_by FROM ac_solicitacoes WHERE id=?", [sol_id]
            )
            if sol_rows:
                cb = json.loads(sol_rows[0].get("created_by") or "{}")
                solicitante_email = cb.get("email", "")
                solicitante_nome  = cb.get("name", "Solicitante")
                if solicitante_email:
                    empresa  = data.get("empresa", "—")
                    cnpj     = data.get("cnpj", "")
                    st_lbl   = _STATUS_LABEL.get(decisao_payload["status"] or "", decisao_payload["status"] or "—")
                    color    = _STATUS_COLOR.get(decisao_payload["status"] or "", "#6366f1")
                    now_str  = datetime.now().strftime("%d/%m/%Y às %H:%M")
                    subject  = f"[Análise Concluída] {empresa} — {st_lbl}"
                    html     = _email_html(f"Análise de Crédito: {st_lbl}", color, [
                        ("Empresa",   empresa),
                        ("CNPJ",      cnpj),
                        ("Status",    st_lbl),
                        ("Analista",  current_user.get("name", "—")),
                        ("Limite",    decisao_payload.get("limiteAprovado") or ""),
                        ("Prazo",     (decisao_payload.get("prazoAprovado") + " dias") if decisao_payload.get("prazoAprovado") else ""),
                        ("Data/hora", now_str),
                    ])
                    asyncio.create_task(_send_email(
                        subject, html, [solicitante_email],
                        from_name=current_user.get("name", ""),
                        from_email=current_user.get("email", ""),
                    ))
        except Exception:
            pass  # não bloqueia a resposta se o e-mail falhar

    return {"ok": True, "decisao_at": decisao_at}


# ── Upload de documentos financeiros ────────────────────────────────────────

_MIME_MAP = {
    ".pdf":  "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc":  "application/msword",
    ".png":  "image/png",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
}

@app.post("/api/docs/{sol_id}/upload")
async def upload_docs(
    sol_id:    str,
    balanco:   List[UploadFile] = File(default=[]),
    contrato:  List[UploadFile] = File(default=[]),
    dre:       List[UploadFile] = File(default=[]),
    fat:       List[UploadFile] = File(default=[]),
    current_user=Depends(_get_current_user),
):
    """Salva documentos financeiros no Turso (persistente no Vercel)."""
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "sol_id inválido")
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    sal = []
    for tipo, uploads in [("balanco", balanco), ("contrato", contrato), ("dre", dre), ("fat", fat)]:
        for f in uploads:
            raw = await f.read()
            if not raw:
                continue
            fname = Path(f.filename or "doc").name
            ext   = Path(fname).suffix.lower()
            mime  = _MIME_MAP.get(ext, f.content_type or "application/octet-stream")
            b64   = base64.standard_b64encode(raw).decode()
            doc_id = f"{sol_id}__{tipo}__{fname}"
            now    = datetime.utcnow().isoformat()
            await _turso_exec(
                "INSERT INTO ac_documents (id, sol_id, tipo, nome, content, mime, size_bytes, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)"
                " ON CONFLICT (id) DO UPDATE SET sol_id=EXCLUDED.sol_id, tipo=EXCLUDED.tipo,"
                " nome=EXCLUDED.nome, content=EXCLUDED.content, mime=EXCLUDED.mime,"
                " size_bytes=EXCLUDED.size_bytes, created_at=EXCLUDED.created_at",
                [doc_id, sol_id, tipo, fname, b64, mime, len(raw), now],
            )
            sal.append(f"{tipo}/{fname}")
    return {"saved": sal, "sol_id": sol_id}


@app.get("/api/docs/{sol_id}/{tipo}/{fname}")
async def download_doc(sol_id: str, tipo: str, fname: str, current_user=Depends(_get_current_user)):
    """Baixa um documento armazenado no Turso."""
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")
    rows = await _turso_query(
        "SELECT content, mime, nome FROM ac_documents WHERE sol_id=? AND tipo=? AND nome=?",
        [sol_id, tipo, fname],
    )
    if not rows:
        raise HTTPException(404, "Arquivo não encontrado.")
    row = rows[0]
    raw = base64.standard_b64decode(row["content"])
    safe_name = Path(row["nome"]).name
    return Response(
        content=raw,
        media_type=row["mime"] or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@app.delete("/api/docs/{sol_id}/{nome}")
@limiter.limit("20/minute")
async def delete_doc(sol_id: str, nome: str, request: Request, current_user=Depends(_get_current_user)):
    """Remove um documento armazenado no Turso pelo nome."""
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "sol_id inválido.")
    rows = await _turso_query(
        "SELECT id FROM ac_documents WHERE sol_id=? AND nome=?",
        [sol_id, nome],
    )
    if not rows:
        raise HTTPException(404, "Documento não encontrado.")
    await _turso_exec(
        "DELETE FROM ac_documents WHERE sol_id=? AND nome=?",
        [sol_id, nome],
    )
    return {"ok": True}


@app.post("/api/idwall/{sol_id}")
@limiter.limit("10/minute")
async def run_idwall_bgc(sol_id: str, request: Request, current_user=Depends(_get_current_user)):
    """Cria relatório BGC PJ na IDwall e retorna protocolo para polling pelo cliente."""
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "sol_id inválido.")

    body = await request.json()
    cnpj = re.sub(r"\D", "", str(body.get("cnpj", "")))
    if len(cnpj) != 14:
        raise HTTPException(400, "CNPJ inválido — informe 14 dígitos.")

    token = os.getenv("IDWALL_API_TOKEN", "")
    if not token:
        raise HTTPException(503, "IDwall não configurado no servidor.")

    hdrs = {"Authorization": token, "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            create_res = await client.post(
                "https://api-v2.idwall.co/relatorios",
                json={"matriz": "vendemmia_bgc_completo_v2_pj", "parametros": {"cnpj_numero": cnpj}},
                headers=hdrs,
            )
    except httpx.TimeoutException:
        raise HTTPException(502, "IDwall: timeout ao conectar — verifique conectividade do servidor.")
    except httpx.RequestError as exc:
        raise HTTPException(502, f"IDwall: erro de rede — {type(exc).__name__}: {str(exc)[:300]}")

    if create_res.status_code not in (200, 201):
        raise HTTPException(502, f"IDwall HTTP {create_res.status_code} ao criar relatório.")

    try:
        result_obj = create_res.json().get("result", {})
    except Exception:
        raise HTTPException(502, "IDwall retornou resposta não-JSON ao criar relatório.")

    # IDwall usa "numero" (não "protocolo") como identificador do relatório
    protocolo = result_obj.get("numero") or result_obj.get("protocolo")
    if not protocolo:
        campos = list(result_obj.keys())
        raise HTTPException(502, f"IDwall não retornou número do relatório. Campos: {campos}")

    # Retorna protocolo imediatamente — polling é feito pelo cliente via /api/idwall-poll
    # debug_fields incluído para diagnóstico durante integração
    return {
        "protocolo": protocolo,
        "status": "EM_EXECUCAO",
        "debug_create_fields": list(result_obj.keys()),
        "debug_create_status": result_obj.get("status", ""),
        "debug_create_msg": result_obj.get("mensagem", ""),
    }


@app.get("/api/idwall-poll/{protocolo}")
@limiter.limit("30/minute")
async def poll_idwall_bgc(protocolo: str, request: Request, cnpj: str = "", current_user=Depends(_get_current_user)):
    """Consulta o status de um relatório IDwall. Retorna dados completos quando CONCLUIDO."""
    if not re.match(r'^[a-zA-Z0-9_\-]{4,128}$', protocolo):
        raise HTTPException(400, "protocolo inválido.")

    token = os.getenv("IDWALL_API_TOKEN", "")
    if not token:
        raise HTTPException(503, "IDwall não configurado no servidor.")

    hdrs = {"Authorization": token, "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            poll = await client.get(
                f"https://api-v2.idwall.co/relatorios/{protocolo}",
                headers=hdrs,
            )
    except httpx.TimeoutException:
        raise HTTPException(502, "IDwall: timeout ao consultar relatório.")
    except httpx.RequestError as exc:
        raise HTTPException(502, f"IDwall: erro de rede — {type(exc).__name__}: {str(exc)[:200]}")

    if poll.status_code != 200:
        raise HTTPException(502, f"IDwall poll HTTP {poll.status_code}.")

    try:
        data = poll.json().get("result", {})
    except Exception:
        raise HTTPException(502, "IDwall retornou resposta não-JSON no poll.")

    status = data.get("status", "")

    if status not in ("CONCLUIDO", "APROVADO", "REPROVADO", "INCONCLUSIVO", "ERRO", "INVALID"):
        return {"status": status, "protocolo": protocolo}

    # Busca dados estruturados dos sub-endpoints /validacoes e /dados
    validacoes: list = []
    dados_extras: dict = {}

    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r_val = await client.get(
                f"https://api-v2.idwall.co/relatorios/{protocolo}/validacoes",
                headers=hdrs,
            )
            if r_val.status_code == 200:
                for v in (r_val.json().get("result", {}).get("validacoes") or []):
                    validacoes.append({
                        "nome": v.get("nome", ""),
                        "descricao": v.get("descricao", ""),
                        "status": v.get("resultado", ""),
                        "resultado": v.get("resultado", ""),
                        "mensagem": v.get("mensagem", ""),
                    })

            r_dados = await client.get(
                f"https://api-v2.idwall.co/relatorios/{protocolo}/dados",
                headers=hdrs,
            )
            if r_dados.status_code == 200:
                d = r_dados.json().get("result", {})
                ci = d.get("cnpj") or {}
                proc = d.get("processos") or {}
                trf = d.get("processos_trf") or {}
                divida = d.get("divida_ativa") or {}
                contrib_list = (d.get("contribuinte") or {}).get("itens") or []
                contrib = contrib_list[0] if contrib_list else {}
                dados_extras = {
                    "empresa": {
                        "razao_social": ci.get("nome_empresarial", ""),
                        "atividade_principal": ci.get("atividade_principal", ""),
                        "capital_social": ci.get("capital_social"),
                        "data_abertura": ci.get("data_abertura", ""),
                        "situacao_cadastral": ci.get("situacao_cadastral", ""),
                        "natureza_juridica": ci.get("natureza_juridica", ""),
                        "cidade": (ci.get("localizacao") or {}).get("cidade", ""),
                        "estado_uf": (ci.get("localizacao") or {}).get("estado", ""),
                        "email": ci.get("email", ""),
                        "qsa": [s.get("nome", "") for s in (ci.get("qsa") or []) if s.get("nome")],
                    },
                    "processos": {
                        "total": len(proc.get("itens") or []),
                        "estados_com_processos": proc.get("estados_com_processos") or [],
                    },
                    "processos_trf": {
                        "total": trf.get("quantidade_processos") or 0,
                        "estados_consultados": len(trf.get("estados_consultados") or []),
                    },
                    "divida_ativa": {
                        "nome": divida.get("nome", ""),
                        "valor_devido": divida.get("valor_devido", ""),
                    } if divida.get("nome") else {},
                    "contribuinte": {
                        "situacao_ie": contrib.get("situacao_ie", ""),
                        "tipo_ie": contrib.get("tipo_ie", ""),
                        "inscricao_estadual": contrib.get("inscricao_estadual", ""),
                        "situacao_documento": contrib.get("situacao_documento", ""),
                        "regime_tributacao": contrib.get("regime_tributacao", ""),
                        "estado": contrib.get("estado", ""),
                        "municipio": contrib.get("municipio", ""),
                    } if contrib else {},
                }
    except Exception:
        pass

    resultado_final = data.get("resultado", "") or status
    mensagem = data.get("mensagem", "")

    return {
        "protocolo": protocolo,
        "status": "CONCLUIDO",
        "resultado": resultado_final,
        "mensagem": mensagem,
        "nomeEmpresa": dados_extras.get("empresa", {}).get("razao_social", ""),
        "cnpj": cnpj,
        "consultadaEm": date.today().isoformat(),
        "validacoes": validacoes,
        "dados": dados_extras,
    }


@app.get("/api/idwall-inspect/{numero}")
@limiter.limit("5/minute")
async def idwall_inspect(numero: str, request: Request, current_user=Depends(_get_current_user)):
    """Diagnóstico: busca todos os sub-endpoints de um relatório para mapear estrutura JSON."""
    if not re.match(r'^[a-zA-Z0-9_\-]{4,128}$', numero):
        raise HTTPException(400, "numero inválido.")
    token = os.getenv("IDWALL_API_TOKEN", "")
    if not token:
        raise HTTPException(503, "IDwall não configurado.")
    hdrs = {"Authorization": token, "Content-Type": "application/json"}
    base = f"https://api-v2.idwall.co/relatorios/{numero}"
    result = {}
    async with httpx.AsyncClient(timeout=15.0) as client:
        for endpoint in ("", "/dados", "/validacoes", "/consultas", "/parametros"):
            try:
                r = await client.get(base + endpoint, headers=hdrs)
                result[endpoint or "/"] = {"status": r.status_code, "body": r.json() if r.status_code == 200 else r.text[:500]}
            except Exception as exc:
                result[endpoint or "/"] = {"error": str(exc)}
    return result


@app.get("/api/idwall-ping")
@limiter.limit("5/minute")
async def idwall_ping(request: Request, current_user=Depends(_get_current_user)):
    """Diagnóstico: testa conectividade e mostra estrutura de um relatório existente."""
    token = os.getenv("IDWALL_API_TOKEN", "")
    token_ok = bool(token)
    token_preview = (token[:8] + "…") if token else "(vazio)"
    hdrs = {"Authorization": token}
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r_list = await client.get("https://api-v2.idwall.co/relatorios", headers=hdrs)
        result = {
            "token_ok": token_ok,
            "token_preview": token_preview,
            "list_status": r_list.status_code,
            "list_preview": r_list.text[:500],
        }
        # Se houver relatórios, busca o primeiro para ver estrutura completa
        itens = r_list.json().get("result", {}).get("itens", []) if r_list.status_code == 200 else []
        if itens:
            numero = itens[0].get("numero")
            async with httpx.AsyncClient(timeout=10.0) as client:
                r_one = await client.get(f"https://api-v2.idwall.co/relatorios/{numero}", headers=hdrs)
            result["single_status"] = r_one.status_code
            result["single_body"] = r_one.text[:1000]
        return result
    except httpx.TimeoutException:
        return {"token_ok": token_ok, "token_preview": token_preview, "error": "timeout"}
    except httpx.RequestError as exc:
        return {"token_ok": token_ok, "token_preview": token_preview, "error": f"{type(exc).__name__}: {str(exc)[:200]}"}


# ── Extração de indicadores financeiros ─────────────────────────────────────

def _xlsx_to_text(data: bytes, filename: str) -> str:
    """Converte Excel para texto tabular."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = [f"=== {filename} ==="]
        for name in wb.sheetnames:
            ws = wb[name]
            lines.append(f"\n--- Planilha: {name} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line  = " | ".join(cells)
                if line.replace("|", "").strip():
                    lines.append(line)
        return "\n".join(lines)
    except Exception as exc:
        return f"[Erro ao ler {filename}: {exc}]"


def _pdf_to_text(data: bytes, filename: str) -> str:
    """Extrai texto e tabelas de PDF usando pdfplumber (sem consumir tokens da IA)."""
    try:
        parts = [f"=== {filename} ==="]
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    parts.append(f"\n--- Página {i} ---\n{text}")
                for table in page.extract_tables():
                    rows = [" | ".join(str(c or "").strip() for c in row) for row in table if any(c for c in row)]
                    if rows:
                        parts.append("\n[Tabela]\n" + "\n".join(rows))
        return "\n".join(parts)
    except Exception as exc:
        return f"[Erro ao ler {filename}: {exc}]"


def _extract_text_from_files(file_list: list) -> str:
    """Extrai texto de todos os arquivos (PDF via pdfplumber, Excel via openpyxl)."""
    parts = []
    for raw, fname in file_list:
        if fname.lower().endswith(".pdf"):
            parts.append(_pdf_to_text(raw, fname))
        elif fname.lower().endswith((".xlsx", ".xls")):
            parts.append(_xlsx_to_text(raw, fname))
    return "\n\n".join(parts)


def _pdf_to_structured(data: bytes, filename: str) -> dict:
    """Extrai conteúdo estruturado (texto + tabelas por página) de um PDF via pdfplumber."""
    try:
        secoes: list[dict] = []
        num_paginas = 0
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            num_paginas = len(pdf.pages)
            for i, page in enumerate(pdf.pages, 1):
                text = (page.extract_text() or "").strip()
                if text:
                    secoes.append({"tipo": "texto", "pagina": i, "conteudo": text})
                for table in page.extract_tables():
                    rows = [
                        [str(c or "").strip() for c in row]
                        for row in table
                        if any(c for c in row)
                    ]
                    if rows:
                        secoes.append({"tipo": "tabela", "pagina": i, "linhas": rows})
        return {"nome": filename, "tipo": "pdf", "paginas": num_paginas, "secoes": secoes}
    except Exception as exc:
        return {"nome": filename, "tipo": "pdf", "paginas": 0, "secoes": [], "erro": str(exc)}


def _xlsx_to_structured(data: bytes, filename: str) -> dict:
    """Extrai conteúdo estruturado (planilhas como tabelas) de um Excel via openpyxl."""
    try:
        secoes: list[dict] = []
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            linhas = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    linhas.append(cells)
            if linhas:
                secoes.append({"tipo": "planilha", "nome": name, "linhas": linhas})
        return {"nome": filename, "tipo": "excel", "planilhas": len(wb.sheetnames), "secoes": secoes}
    except Exception as exc:
        return {"nome": filename, "tipo": "excel", "planilhas": 0, "secoes": [], "erro": str(exc)}


def _extract_documents_structured(file_list: list) -> list:
    """Retorna lista de documentos com conteúdo estruturado (sem IA)."""
    docs = []
    for raw, fname in file_list:
        if fname.lower().endswith(".pdf"):
            docs.append(_pdf_to_structured(raw, fname))
        elif fname.lower().endswith((".xlsx", ".xls")):
            docs.append(_xlsx_to_structured(raw, fname))
    return docs


@app.post("/api/extract-financials")
@limiter.limit("10/minute")
async def extract_financials(
    request: Request,
    empresa: str = Form(""),
    cnpj:    str = Form(""),
    sol_id:  str = Form(""),
    files:   List[UploadFile] = File(default=[]),
    current_user=Depends(_get_current_user),
):
    """Extrai conteúdo de documentos financeiros via Python (pdfplumber/openpyxl). Sem IA.

    Aceita duas fontes (em ordem de prioridade):
    1. sol_id — lê arquivos já salvos no Turso (tipo IN 'balanco','dre')
    2. files  — upload direto de arquivos pelo usuário
    """
    file_list: list[tuple[bytes, str]] = []

    # Prioridade 1: arquivos salvos no Turso
    if sol_id and _SOL_ID_RE.match(sol_id) and _turso_ok():
        rows = await _turso_query(
            "SELECT nome, content FROM ac_documents WHERE sol_id=? AND tipo IN ('balanco','dre') ORDER BY tipo, nome",
            [sol_id],
        )
        for row in rows:
            ext = Path(row["nome"]).suffix.lower()
            if ext in (".pdf", ".xlsx", ".xls"):
                raw = base64.standard_b64decode(row["content"])
                file_list.append((raw, row["nome"]))

    # Prioridade 2: upload direto (fallback)
    if not file_list and files:
        for f in files:
            raw = await f.read()
            if raw:
                file_list.append((raw, f.filename or "documento"))

    if not file_list:
        raise HTTPException(
            404 if sol_id else 400,
            "Nenhum documento encontrado. Envie os arquivos ou verifique se o upload foi realizado."
        )

    documentos = _extract_documents_structured(file_list)

    total_paginas = sum(d.get("paginas", 0) for d in documentos)
    total_tabelas = sum(
        sum(1 for s in d.get("secoes", []) if s["tipo"] in ("tabela", "planilha"))
        for d in documentos
    )
    nd = len(documentos)
    resumo = f"{nd} documento{'s' if nd != 1 else ''} analisado{'s' if nd != 1 else ''}"
    if total_paginas:
        resumo += f" • {total_paginas} página{'s' if total_paginas != 1 else ''}"
    if total_tabelas:
        resumo += f" • {total_tabelas} tabela{'s' if total_tabelas != 1 else ''} encontrada{'s' if total_tabelas != 1 else ''}"

    return {"documentos": documentos, "resumo": resumo}


# ── Análise Contábil ──────────────────────────────────────────────────────────

class PeriodoBP(BaseModel):
    disponibilidade:      Optional[float] = None
    contas_receber:       Optional[float] = None
    estoques:             Optional[float] = None
    impostos_recuperar:   Optional[float] = None
    outros_ac:            Optional[float] = None
    outros_creditos:      Optional[float] = None
    imobilizado:          Optional[float] = None
    investimentos:        Optional[float] = None
    outros_anc:           Optional[float] = None
    fornecedores:         Optional[float] = None
    adiantamento_clientes: Optional[float] = None
    impostos_pagar:       Optional[float] = None
    emprestimos_cp:       Optional[float] = None
    outros_pc:            Optional[float] = None
    emprestimos_lp:       Optional[float] = None
    outros_pnc:           Optional[float] = None
    patrimonio_liquido:   Optional[float] = None


class PeriodoDRE(BaseModel):
    receita_bruta:          Optional[float] = None
    deducoes:               Optional[float] = None
    receita_liquida:        Optional[float] = None
    cpv:                    Optional[float] = None
    lucro_bruto:            Optional[float] = None
    despesas_operacionais:  Optional[float] = None
    ebitda:                 Optional[float] = None
    resultado_financeiro:   Optional[float] = None
    lucro_antes_ir:         Optional[float] = None
    ir_csll:                Optional[float] = None
    lucro_liquido:          Optional[float] = None


class ContabilData(BaseModel):
    periodo1_label:    str = ""
    periodo2_label:    str = ""
    bp1:               PeriodoBP  = PeriodoBP()
    bp2:               PeriodoBP  = PeriodoBP()
    dre1:              PeriodoDRE = PeriodoDRE()
    dre2:              PeriodoDRE = PeriodoDRE()
    fator_risco:       float = 0.10
    fator_multiplicador: float = 1.1
    cdi:               float = 0.149  # CDI anual (default 14,9% aa — atualize conforme mercado)


def _sdiv(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def _sem_liq_seca(v):
    if v is None: return "nd"
    return "ok" if v >= 1.0 else ("warn" if v >= 0.7 else "bad")

def _sem_liq_geral(v):
    if v is None: return "nd"
    return "ok" if v >= 2.0 else ("warn" if v >= 1.0 else "bad")

def _sem_roe(v):
    if v is None: return "nd"
    return "ok" if v >= 0.10 else ("warn" if v >= 0.05 else "bad")

def _sem_roa(v):
    if v is None: return "nd"
    return "ok" if v >= 0.05 else ("warn" if v >= 0.02 else "bad")

def _sem_endiv(v):
    if v is None: return "nd"
    return "ok" if v <= 0.50 else ("warn" if v <= 1.00 else "bad")

def _sem_margem_bruta(v):
    if v is None: return "nd"
    return "ok" if v >= 0.25 else ("warn" if v >= 0.15 else "bad")

def _sem_margem_liq(v):
    if v is None: return "nd"
    return "ok" if v >= 0.05 else ("warn" if v >= 0.01 else "bad")

def _sem_pmr(v):
    if v is None: return "nd"
    return "ok" if v <= 60 else ("warn" if v <= 120 else "bad")

def _sem_pme(v):
    if v is None: return "nd"
    return "ok" if v <= 60 else ("warn" if v <= 120 else "bad")


def _calc_periodo(bp: PeriodoBP, dre: PeriodoDRE, fator_risco: float, fator_mult: float, cdi: float = 0.149) -> dict:
    def nn(*vals):
        return sum(v for v in vals if v is not None)

    ac  = nn(bp.disponibilidade, bp.contas_receber, bp.estoques, bp.impostos_recuperar, bp.outros_ac)
    anc = nn(bp.outros_creditos, bp.imobilizado, bp.investimentos, bp.outros_anc)
    at  = ac + anc if (bp.disponibilidade is not None or bp.imobilizado is not None) else None
    pc  = nn(bp.fornecedores, bp.adiantamento_clientes, bp.impostos_pagar, bp.emprestimos_cp, bp.outros_pc)
    pnc = nn(bp.emprestimos_lp, bp.outros_pnc)
    pl  = bp.patrimonio_liquido
    passivo_total = pc + pnc if (bp.fornecedores is not None or bp.emprestimos_lp is not None) else None

    # DRE: derivar RL e LB se não fornecidos diretamente
    rl = dre.receita_liquida
    if rl is None and dre.receita_bruta is not None:
        rl = dre.receita_bruta - abs(dre.deducoes or 0)
    lb = dre.lucro_bruto
    if lb is None and rl is not None and dre.cpv is not None:
        lb = rl - abs(dre.cpv)
    ll  = dre.lucro_liquido
    cpv = dre.cpv

    liq_seca     = _sdiv((ac - (bp.estoques or 0)) if at is not None else None, pc if pc else None)
    liq_geral    = _sdiv(at, passivo_total)
    roe          = _sdiv(ll, pl)
    roa          = _sdiv(ll, at)
    margem_bruta = _sdiv(lb, rl)
    margem_liq   = _sdiv(ll, rl)

    ativ_op = nn(bp.contas_receber, bp.estoques, bp.outros_ac)
    pass_op = nn(bp.fornecedores, bp.adiantamento_clientes, bp.outros_pc)
    ncg = ativ_op - pass_op if (bp.contas_receber is not None or bp.fornecedores is not None) else None

    endiv_geral   = _sdiv(passivo_total, at)
    emp_total     = (bp.emprestimos_cp or 0) + (bp.emprestimos_lp or 0)
    endiv_oneroso = _sdiv(emp_total if (bp.emprestimos_cp or bp.emprestimos_lp) else None, pl)

    pmr = (bp.contas_receber / rl * 360) if (bp.contas_receber and rl) else None
    pme = (bp.estoques / abs(cpv) * 360)  if (bp.estoques and cpv) else None

    # ── Cálculo de limite — 4 passos (metodologia estudo.xlsx) ────────────────
    # Passo 1: Capacidade de pagamento LP = Liq. Geral × PL
    cap_pagamento_lp = (liq_geral * pl) if (liq_geral and pl) else None

    # Passo 2: Capacidade de retorno = PL × ROE × (1 + CDI)
    cap_retorno = (pl * roe * (1 + cdi)) if (pl and roe is not None) else None

    # Passo 3: Fator de cobertura = Cap. Retorno / Lucro Líquido
    fator_cobertura = _sdiv(cap_retorno, ll)

    # Passo 4: Crédito proposto = Cap. Retorno × Fator Cobertura × (1 - Fator Risco)
    if cap_retorno is not None and cap_retorno > 0 and fator_cobertura is not None:
        credito_proposto = cap_retorno * fator_cobertura * (1 - fator_risco)
    elif cap_retorno is not None and cap_retorno <= 0:
        credito_proposto = 0.0
    else:
        # Fallback quando LL é nulo: usa LL × mult / risco
        credito_proposto = (ll * fator_mult / fator_risco) if (ll and fator_risco) else None

    credito_calculado = cap_pagamento_lp  # Passo 1 = teto máximo pela liquidez

    def r(v, dec=2):
        return round(v, dec) if v is not None else None

    def ind(label, formula, valor, status, valor_pct=None, unidade=None):
        return {
            "label": label, "formula": formula,
            "valor": r(valor), "valor_pct": r(valor_pct, 2),
            "status": status, "unidade": unidade,
        }

    return {
        "totais": {
            "ativo_circulante":    r(ac, 0)  if at is not None else None,
            "ativo_nao_circulante": r(anc, 0) if at is not None else None,
            "ativo_total":         r(at, 0),
            "passivo_circulante":  r(pc, 0)  if passivo_total is not None else None,
            "passivo_nao_circulante": r(pnc, 0) if passivo_total is not None else None,
            "passivo_total":       r(passivo_total, 0),
            "patrimonio_liquido":  r(pl, 0),
            "receita_liquida":     r(rl, 0),
            "lucro_bruto":         r(lb, 0),
            "lucro_liquido":       r(ll, 0),
        },
        "indicadores": {
            "margem_bruta":        ind("Margem Bruta",        "LB ÷ RL",           margem_bruta,  _sem_margem_bruta(margem_bruta), valor_pct=round(margem_bruta*100,1) if margem_bruta is not None else None, unidade="%"),
            "margem_liquida":      ind("Margem Líquida",      "LL ÷ RL",           margem_liq,    _sem_margem_liq(margem_liq),   valor_pct=round(margem_liq*100,1)   if margem_liq   is not None else None, unidade="%"),
            "liquidez_seca":       ind("Liquidez Seca",       "(AC−Est) ÷ PC",     liq_seca,      _sem_liq_seca(liq_seca),       unidade="×"),
            "liquidez_geral":      ind("Liquidez Geral",      "AT ÷ (PC+PNC)",     liq_geral,     _sem_liq_geral(liq_geral),     unidade="×"),
            "roe":                 ind("ROE",                 "LL ÷ PL",           roe,           _sem_roe(roe),                 valor_pct=round(roe*100,1)           if roe is not None else None, unidade="%"),
            "roa":                 ind("ROA",                 "LL ÷ AT",           roa,           _sem_roa(roa),                 valor_pct=round(roa*100,1)           if roa is not None else None, unidade="%"),
            "ncg":                 ind("NCG",                 "(CR+Est+OAC)−(Forn+Adiant+OPC)", ncg, "ok" if ncg is not None and ncg >= 0 else ("bad" if ncg is not None else "nd"), unidade="R$"),
            "endividamento_geral": ind("Endividamento Geral", "(PC+PNC) ÷ AT",     endiv_geral,   _sem_endiv(endiv_geral),       valor_pct=round(endiv_geral*100,1)   if endiv_geral is not None else None, unidade="%"),
            "endividamento_oneroso": ind("Endiv. Oneroso",   "(EmpCP+EmpLP) ÷ PL",endiv_oneroso, _sem_endiv(endiv_oneroso),     valor_pct=round(endiv_oneroso*100,1) if endiv_oneroso is not None else None, unidade="%"),
            "pmr":                 ind("PMR",                 "(CR ÷ RL) × 360",   pmr,           _sem_pmr(pmr),                 unidade="dias"),
            "pme":                 ind("PME",                 "(Est ÷ CPV) × 360", pme,           _sem_pme(pme),                 unidade="dias"),
        },
        "credito": {
            "calculado":         r(credito_calculado, 0),
            "proposto":          r(credito_proposto, 0),
            "cap_pagamento_lp":  r(cap_pagamento_lp, 0),
            "cap_retorno":       r(cap_retorno, 0),
            "fator_cobertura":   r(fator_cobertura, 4),
            "cdi":               cdi,
            "fator_risco":       fator_risco,
            "fator_multiplicador": fator_mult,
        },
    }


def _build_contabil_section(contabil_result: dict) -> str:
    """Formata os indicadores contábeis para injeção no prompt do Gemini."""
    p1  = contabil_result.get("periodo1") or {}
    p2  = contabil_result.get("periodo2") or {}
    l1  = contabil_result.get("periodo1_label") or "Período 1"
    l2  = contabil_result.get("periodo2_label") or "Período 2"
    t2  = p2.get("totais") or {}
    i1  = p1.get("indicadores") or {}
    i2  = p2.get("indicadores") or {}
    cr2 = p2.get("credito") or {}

    sem_map = {"ok": "✔", "warn": "⚠", "bad": "✗", "nd": "—"}

    def fmtbr(v):
        if v is None: return "—"
        return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

    rows = []
    for k, (lbl, use_pct) in {
        "margem_bruta":        ("Margem Bruta", True),
        "margem_liquida":      ("Margem Líquida", True),
        "liquidez_seca":       ("Liquidez Seca", False),
        "liquidez_geral":      ("Liquidez Geral", False),
        "roe":                 ("ROE", True),
        "roa":                 ("ROA", True),
        "endividamento_geral": ("Endividamento Geral", True),
        "ncg":                 ("NCG", False),
        "pmr":                 ("PMR (dias)", False),
        "pme":                 ("PME (dias)", False),
    }.items():
        iv1 = i1.get(k) or {}
        iv2 = i2.get(k) or {}
        def _fmtv(iv, pct):
            vp = iv.get("valor_pct")
            v  = iv.get("valor")
            u  = iv.get("unidade", "")
            if pct and vp is not None: return f"{vp}%"
            if v is not None: return f"{v}{' '+u if u not in ('%','') else ''}"
            return "—"
        sem = sem_map.get(iv2.get("status", "nd"), "—")
        rows.append(f"  - {lbl}: {_fmtv(iv1, use_pct)} → {_fmtv(iv2, use_pct)} {sem}")

    return f"""

## ANÁLISE CONTÁBIL (demonstrativos financeiros)
Períodos analisados: {l1} → {l2}

### Posição Patrimonial em {l2}
- Ativo Total: {fmtbr(t2.get('ativo_total'))}
- Passivo Total: {fmtbr(t2.get('passivo_total'))}
- Patrimônio Líquido: {fmtbr(t2.get('patrimonio_liquido'))}
- Receita Líquida: {fmtbr(t2.get('receita_liquida'))}
- Lucro Líquido: {fmtbr(t2.get('lucro_liquido'))}

### Evolução dos Indicadores ({l1} → {l2})
{chr(10).join(rows)}

### Capacidade de Crédito Calculada
- Crédito Base (Liq.Geral × PL): {fmtbr(cr2.get('calculado'))}
- Crédito Proposto (Calculadora Contábil): {fmtbr(cr2.get('proposto'))}

Leve esses indicadores em conta na análise de proporcionalidade e recomendação de limite.
"""


@app.post("/api/contabil/extrair/{sol_id}")
@limiter.limit("10/minute")
async def contabil_extrair(sol_id: str, request: Request, current_user=Depends(_get_current_user)):
    """Extrai dados contábeis estruturados dos PDFs de BP e DRE (tenta Gemini, fallback Claude)."""
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    if not _turso_ok():
        raise HTTPException(503, "Banco de dados não configurado.")

    rows = await _turso_query(
        "SELECT nome, content FROM ac_documents WHERE sol_id=? AND tipo IN ('balanco','dre') ORDER BY tipo, nome",
        [sol_id],
    )
    if not rows:
        raise HTTPException(404, "Nenhum documento de Balanço ou DRE encontrado. Envie os documentos antes de extrair.")

    partes = []
    for row in rows:
        ext = Path(row["nome"]).suffix.lower()
        raw = base64.standard_b64decode(row["content"])
        doc = _pdf_to_structured(raw, row["nome"]) if ext == ".pdf" else _xlsx_to_structured(raw, row["nome"])
        linhas = [f"=== DOCUMENTO: {row['nome']} ==="]
        for sec in doc.get("secoes", []):
            if sec["tipo"] == "texto":
                linhas.append(sec["conteudo"])
            elif sec["tipo"] in ("tabela", "planilha"):
                for lr in sec.get("linhas", []):
                    linhas.append(" | ".join(str(c) for c in lr))
        partes.append("\n".join(linhas))

    texto = "\n\n".join(partes)
    if not texto.strip():
        raise HTTPException(422, "Não foi possível extrair texto dos documentos financeiros.")

    gemini_prompt = f"""Você é um analista contábil especializado em demonstrações financeiras brasileiras.

Analise os documentos abaixo e extraia os dados de Balanço Patrimonial (BP) e DRE para DOIS períodos distintos.
São obrigatoriamente dois períodos (ex: dez/2024 e jun/2025, ou 2023 e 2024).

REGRAS:
- Valores em REAIS, número puro (float), sem formatação
- Custos e despesas: valores POSITIVOS (o sistema aplica o sinal)
- Campo inexistente: null
- Se só encontrar um período, use null em todo o segundo período

DOCUMENTOS:
{texto[:18000]}

Retorne APENAS um JSON válido (sem markdown, sem texto extra):

{{
  "periodo1_label": "<ex: dez/2024>",
  "periodo2_label": "<ex: jun/2025>",
  "bp1": {{
    "disponibilidade": <float|null>,
    "contas_receber": <float|null>,
    "estoques": <float|null>,
    "impostos_recuperar": <float|null>,
    "outros_ac": <float|null>,
    "outros_creditos": <float|null>,
    "imobilizado": <float|null>,
    "investimentos": <float|null>,
    "outros_anc": <float|null>,
    "fornecedores": <float|null>,
    "adiantamento_clientes": <float|null>,
    "impostos_pagar": <float|null>,
    "emprestimos_cp": <float|null>,
    "outros_pc": <float|null>,
    "emprestimos_lp": <float|null>,
    "outros_pnc": <float|null>,
    "patrimonio_liquido": <float|null>
  }},
  "bp2": {{ (mesma estrutura) }},
  "dre1": {{
    "receita_bruta": <float|null>,
    "deducoes": <float|null>,
    "receita_liquida": <float|null>,
    "cpv": <float|null>,
    "lucro_bruto": <float|null>,
    "despesas_operacionais": <float|null>,
    "ebitda": <float|null>,
    "resultado_financeiro": <float|null>,
    "lucro_antes_ir": <float|null>,
    "ir_csll": <float|null>,
    "lucro_liquido": <float|null>
  }},
  "dre2": {{ (mesma estrutura) }},
  "fator_risco": 0.10,
  "fator_multiplicador": 1.1
}}"""

    extracted = None
    ai_error_msg = None

    # ── Tentativa 1: Gemini ───────────────────────────────────────────────────
    gemini_key = _load_key()
    if gemini_key:
        try:
            raw_resp = await asyncio.to_thread(_gemini_generate, gemini_key, gemini_prompt)
            extracted = _extract_json(raw_resp)
        except Exception as exc:
            ai_error_msg = f"Gemini: {str(exc)[:200]}"

    # ── Tentativa 2: Claude (Anthropic) — fallback automático ─────────────────
    if not extracted:
        anthropic_key = _load_anthropic_key()
        if anthropic_key:
            try:
                # Reutiliza o mesmo schema do prompt Gemini — Claude entende igual
                schema_prompt = """Você é um analista contábil especializado em demonstrações financeiras brasileiras.

Analise os documentos abaixo e extraia os dados de Balanço Patrimonial (BP) e DRE para DOIS períodos distintos.
São obrigatoriamente dois períodos (ex: dez/2024 e jun/2025, ou 2023 e 2024).

REGRAS:
- Valores em REAIS, número puro (float), sem formatação
- Custos e despesas: valores POSITIVOS (o sistema aplica o sinal)
- Campo inexistente: null
- Se só encontrar um período, use null em todo o segundo período

Retorne APENAS um JSON válido (sem markdown, sem texto extra):

{
  "periodo1_label": "<ex: dez/2024>",
  "periodo2_label": "<ex: jun/2025>",
  "bp1": {
    "disponibilidade": null, "contas_receber": null, "estoques": null,
    "impostos_recuperar": null, "outros_ac": null, "outros_creditos": null,
    "imobilizado": null, "investimentos": null, "outros_anc": null,
    "fornecedores": null, "adiantamento_clientes": null, "impostos_pagar": null,
    "emprestimos_cp": null, "outros_pc": null, "emprestimos_lp": null,
    "outros_pnc": null, "patrimonio_liquido": null
  },
  "bp2": { (mesma estrutura) },
  "dre1": {
    "receita_bruta": null, "deducoes": null, "receita_liquida": null,
    "cpv": null, "lucro_bruto": null, "despesas_operacionais": null,
    "ebitda": null, "resultado_financeiro": null, "lucro_antes_ir": null,
    "ir_csll": null, "lucro_liquido": null
  },
  "dre2": { (mesma estrutura) },
  "fator_risco": 0.10,
  "fator_multiplicador": 1.1
}"""
                extracted = await asyncio.to_thread(
                    _claude_extract_json, texto, schema_prompt, anthropic_key
                )
                if extracted:
                    ai_error_msg = None  # Claude funcionou — limpa erro anterior
            except Exception as exc:
                ai_error_msg = (ai_error_msg or "") + f" | Claude: {str(exc)[:200]}"

    # ── Tentativa 3: Extração por regex — sem IA, sem chave ──────────────────
    if not extracted:
        try:
            extracted = _extract_contabil_regex(texto)
            if extracted:
                ai_error_msg = None  # regex funcionou
        except Exception as exc:
            ai_error_msg = (ai_error_msg or "") + f" | Regex: {str(exc)[:100]}"

    # ── Fallback final: formulário vazio para preenchimento manual ─────────────
    if not extracted:
        return {
            "ok": True,
            "data": {},
            "result": {},
            "needs_manual": True,
            "texto_extraido": texto[:3000] if texto else "",
            "ai_error": ai_error_msg or "Não foi possível extrair dados automaticamente",
        }

    # Calcular indicadores imediatamente
    try:
        cd = ContabilData(**extracted)
        p1 = _calc_periodo(cd.bp1, cd.dre1, cd.fator_risco, cd.fator_multiplicador, cd.cdi)
        p2 = _calc_periodo(cd.bp2, cd.dre2, cd.fator_risco, cd.fator_multiplicador, cd.cdi)
        contabil_result = {
            "periodo1_label": cd.periodo1_label,
            "periodo2_label": cd.periodo2_label,
            "periodo1": p1,
            "periodo2": p2,
            "fator_risco": cd.fator_risco,
            "fator_multiplicador": cd.fator_multiplicador,
            "raw": extracted,
        }
    except Exception:
        contabil_result = {"raw": extracted}

    # Salvar na solicitação
    try:
        sol_rows = await _turso_query("SELECT data FROM ac_solicitacoes WHERE id=?", [sol_id])
        if sol_rows:
            sol_data = json.loads(sol_rows[0]["data"] or "{}")
            sol_data["contabil_data"] = extracted
            sol_data["contabil_result"] = contabil_result
            sol_data["contabil_extraido_at"] = datetime.utcnow().isoformat()
            await _turso_exec(
                "UPDATE ac_solicitacoes SET data=?, updated_at=? WHERE id=?",
                [json.dumps(sol_data, ensure_ascii=False), datetime.utcnow().isoformat(), sol_id],
            )
    except Exception:
        pass

    return {"ok": True, "data": extracted, "result": contabil_result}


@app.post("/api/contabil/calcular")
@limiter.limit("60/minute")
async def contabil_calcular(request: Request, current_user=Depends(_get_current_user)):
    """Recalcula todos os indicadores a partir de dados de BP/DRE editados pelo analista."""
    body = await request.json()
    try:
        raw_data = body.get("data", body)
        cd = ContabilData(**raw_data)
    except Exception as exc:
        raise HTTPException(400, f"Dados inválidos: {exc}")

    p1 = _calc_periodo(cd.bp1, cd.dre1, cd.fator_risco, cd.fator_multiplicador, cd.cdi)
    p2 = _calc_periodo(cd.bp2, cd.dre2, cd.fator_risco, cd.fator_multiplicador, cd.cdi)

    return {
        "periodo1_label":    cd.periodo1_label,
        "periodo2_label":    cd.periodo2_label,
        "periodo1":          p1,
        "periodo2":          p2,
        "fator_risco":       cd.fator_risco,
        "fator_multiplicador": cd.fator_multiplicador,
        "cdi":               cd.cdi,
    }


@app.patch("/api/contabil/{sol_id}")
@limiter.limit("60/minute")
async def contabil_salvar(sol_id: str, request: Request, current_user=Depends(_get_current_user)):
    """Persiste alterações manuais do analista nos dados contábeis."""
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    if not _turso_ok():
        raise HTTPException(503, "Banco não configurado.")
    patch = await request.json()
    sol_rows = await _turso_query("SELECT data FROM ac_solicitacoes WHERE id=?", [sol_id])
    if not sol_rows:
        raise HTTPException(404, "Solicitação não encontrada")
    sol_data = json.loads(sol_rows[0]["data"] or "{}")
    existing = sol_data.get("contabil_data") or {}
    existing.update(patch)
    # Recalcular após patch
    try:
        cd = ContabilData(**existing)
        p1 = _calc_periodo(cd.bp1, cd.dre1, cd.fator_risco, cd.fator_multiplicador, cd.cdi)
        p2 = _calc_periodo(cd.bp2, cd.dre2, cd.fator_risco, cd.fator_multiplicador, cd.cdi)
        sol_data["contabil_result"] = {
            "periodo1_label": cd.periodo1_label,
            "periodo2_label": cd.periodo2_label,
            "periodo1": p1, "periodo2": p2,
            "fator_risco": cd.fator_risco,
            "fator_multiplicador": cd.fator_multiplicador,
        }
    except Exception:
        pass
    sol_data["contabil_data"]    = existing
    sol_data["contabil_editado_at"] = datetime.utcnow().isoformat()
    await _turso_exec(
        "UPDATE ac_solicitacoes SET data=?, updated_at=? WHERE id=?",
        [json.dumps(sol_data, ensure_ascii=False), datetime.utcnow().isoformat(), sol_id],
    )
    return {"ok": True, "result": sol_data.get("contabil_result")}


# ── Notificações por e-mail ───────────────────────────────────────────────────

class NotifyEmailRequest(BaseModel):
    event:       str
    empresa:     Optional[str] = ""
    cnpj:        Optional[str] = ""
    solicitante: Optional[str] = ""
    status:      Optional[str] = ""
    limite:      Optional[str] = ""
    prazo:       Optional[str] = ""
    analista:    Optional[str] = ""
    deliberacao: Optional[str] = ""


@app.post("/api/notify/email")
@limiter.limit("30/minute")
async def notify_email(
    request: Request,
    body: NotifyEmailRequest,
    current_user=Depends(_get_current_user),
):
    if not _NOTIFY_EMAILS:
        return {"ok": False, "reason": "NOTIFY_EMAILS não configurado"}

    now_str = datetime.now().strftime("%d/%m/%Y às %H:%M")
    empresa = body.empresa or "—"
    cnpj    = body.cnpj or ""

    if body.event == "nova_solicitacao":
        subject  = f"[Nova Solicitação] {empresa}"
        headline = "Nova Solicitação de Crédito"
        color    = "#6366f1"
        rows = [
            ("Empresa",     empresa),
            ("CNPJ",        cnpj),
            ("Solicitante", body.solicitante or current_user.get("name", "—")),
            ("Data/hora",   now_str),
        ]

    elif body.event == "analista_decisao":
        st_lbl   = _STATUS_LABEL.get(body.status or "", body.status or "—")
        color    = _STATUS_COLOR.get(body.status or "", "#6366f1")
        subject  = f"[{st_lbl}] {empresa} — Decisão do Analista"
        headline = f"Decisão do Analista: {st_lbl}"
        rows = [
            ("Empresa",    empresa),
            ("CNPJ",       cnpj),
            ("Status",     st_lbl),
            ("Analista",   body.analista or current_user.get("name", "—")),
            ("Limite",     body.limite or ""),
            ("Prazo",      (body.prazo + " dias") if body.prazo else ""),
            ("Data/hora",  now_str),
        ]

    elif body.event == "comite_decisao":
        st_lbl   = _STATUS_LABEL.get(body.status or "", body.status or "—")
        color    = _STATUS_COLOR.get(body.status or "", "#f59e0b")
        subject  = f"[Comitê — {st_lbl}] {empresa}"
        headline = f"Decisão do Comitê: {st_lbl}"
        rows = [
            ("Empresa",       empresa),
            ("CNPJ",          cnpj),
            ("Status Final",  st_lbl),
            ("Deliberação",   body.deliberacao or "—"),
            ("Limite Final",  body.limite or ""),
            ("Data/hora",     now_str),
        ]

    else:
        return {"ok": False, "reason": "Evento desconhecido"}

    html = _email_html(headline, color, rows)
    asyncio.create_task(_send_email(
        subject, html, _NOTIFY_EMAILS,
        from_name=current_user.get("name", ""),
        from_email=current_user.get("email", ""),
    ))
    return {"ok": True}


# ── Admin: Gestão de Usuários ─────────────────────────────────────────────────

def _welcome_html_api(name: str, url: str) -> str:
    logo_url = os.getenv("ALLOWED_ORIGINS", "https://analisecredito.vendemmia.dev.br").split(",")[0].strip() + "/logo.png"
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f8;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f8;padding:40px 0;">
  <tr><td align="center">
    <table width="600" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.10);">
      <tr><td style="background:linear-gradient(135deg,#1e1b4b 0%,#312e81 100%);padding:32px 40px;">
        <table width="100%" cellpadding="0" cellspacing="0"><tr>
          <td><img src="{logo_url}" alt="Vendemmia" height="48" style="height:48px;max-width:200px;object-fit:contain;display:block;" /></td>
          <td align="right" style="color:rgba(255,255,255,.5);font-size:11px;letter-spacing:.5px;text-transform:uppercase;vertical-align:bottom;">Análise de Crédito</td>
        </tr></table>
      </td></tr>
      <tr><td style="background:linear-gradient(90deg,#4f46e5,#7c3aed);padding:20px 40px;">
        <p style="margin:0;color:#fff;font-size:20px;font-weight:700;">Bem-vindo ao sistema!</p>
        <p style="margin:6px 0 0;color:rgba(255,255,255,.75);font-size:13px;">Sua conta foi criada. Defina sua senha para começar.</p>
      </td></tr>
      <tr><td style="padding:36px 40px 28px;">
        <p style="margin:0 0 16px;font-size:16px;color:#1e1b4b;font-weight:700;">Olá, {name}!</p>
        <p style="margin:0 0 12px;font-size:14px;color:#555;line-height:1.7;">
          Você foi cadastrado no <strong>Sistema de Análise de Crédito da Vendemmia</strong>.
          Clique no botão abaixo para definir sua senha e ativar seu acesso.
        </p>
        <p style="margin:0 0 28px;font-size:13px;color:#888;">O link é <strong>válido por 24 horas</strong>.</p>
        <table cellpadding="0" cellspacing="0" style="margin-bottom:28px;">
          <tr><td style="border-radius:12px;background:linear-gradient(135deg,#4f46e5,#7c3aed);box-shadow:0 4px 14px rgba(79,70,229,.4);">
            <a href="{url}" style="display:inline-block;padding:16px 40px;color:#fff;text-decoration:none;font-size:15px;font-weight:700;border-radius:12px;">Definir minha senha</a>
          </td></tr>
        </table>
        <p style="margin:0;font-size:12px;word-break:break-all;"><a href="{url}" style="color:#6366f1;">{url}</a></p>
      </td></tr>
      <tr><td style="padding:20px 40px;background:#f9f9fb;border-top:1px solid #ebebeb;text-align:center;">
        <p style="margin:0;font-size:11px;color:#aaa;">Sistema interno Vendemmia &middot; Não responda este e-mail</p>
      </td></tr>
    </table>
  </td></tr>
</table></body></html>"""


async def _create_welcome_token_and_send(name: str, email: str) -> bool:
    await _ensure_reset_tables()
    token      = secrets.token_urlsafe(32)
    expires_at = (datetime.utcnow() + timedelta(hours=24)).isoformat()
    await _turso_exec(
        "INSERT INTO ac_password_reset_tokens (token, email, expires_at, used) VALUES (?,?,?,0)"
        " ON CONFLICT (token) DO NOTHING",
        [token, email, expires_at],
    )
    base_url = os.getenv("ALLOWED_ORIGINS", "http://localhost:8000").split(",")[0].strip()
    link     = f"{base_url}/login.html?reset={token}&welcome=1"
    html     = _welcome_html_api(name, link)
    await _send_email(
        "Bem-vindo ao Sistema de Análise de Crédito — Vendemmia",
        html, [email],
        from_name="Vendemmia Análise de Crédito",
    )
    return True


@app.get("/api/admin/users")
async def admin_list_users(current_user=Depends(_require_admin)):
    rows = await _turso_query(
        "SELECT id, name, email, role, avatar, hashed_password, created_at, updated_at FROM ac_users ORDER BY name"
    )
    return [
        {
            "id":         r["id"],
            "name":       r["name"],
            "email":      r["email"],
            "role":       r["role"],
            "avatar":     r["avatar"],
            "ativo":      bool(r.get("hashed_password")),
            "created_at": r.get("created_at"),
            "updated_at": r.get("updated_at"),
        }
        for r in rows
    ]


class UserCreateRequest(BaseModel):
    name:  str
    email: str
    role:  str


@app.post("/api/admin/users", status_code=201)
async def admin_create_user(body: UserCreateRequest, current_user=Depends(_require_admin)):
    email = body.email.strip().lower()
    name  = body.name.strip()
    role  = body.role.strip()

    existing = await _turso_query("SELECT id FROM ac_users WHERE email=?", [email])
    if existing:
        raise HTTPException(400, "Já existe um usuário com este e-mail.")

    uid     = "u_" + secrets.token_urlsafe(8)
    avatar  = (name[:2]).upper()
    now_iso = datetime.utcnow().isoformat()
    await _turso_exec(
        "INSERT INTO ac_users (id, name, email, hashed_password, role, avatar, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)",
        [uid, name, email, "", role, avatar, now_iso, now_iso],
    )

    email_sent = False
    if _SMTP_HOST:
        try:
            await _create_welcome_token_and_send(name, email)
            email_sent = True
        except Exception:
            pass

    return {"ok": True, "id": uid, "email_sent": email_sent}


class UserUpdateRequest(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None


@app.put("/api/admin/users/{user_id}")
async def admin_update_user(user_id: str, body: UserUpdateRequest, current_user=Depends(_require_admin)):
    rows = await _turso_query("SELECT id, name, role, avatar FROM ac_users WHERE id=?", [user_id])
    if not rows:
        raise HTTPException(404, "Usuário não encontrado.")
    u       = rows[0]
    name    = (body.name or u["name"]).strip()
    role    = (body.role or u["role"]).strip()
    avatar  = (name[:2]).upper()
    now_iso = datetime.utcnow().isoformat()
    await _turso_exec(
        "UPDATE ac_users SET name=?, role=?, avatar=?, updated_at=? WHERE id=?",
        [name, role, avatar, now_iso, user_id],
    )
    return {"ok": True}


@app.delete("/api/admin/users/{user_id}", status_code=204)
async def admin_delete_user(user_id: str, current_user=Depends(_require_admin)):
    if current_user.get("sub") == user_id:
        raise HTTPException(400, "Você não pode excluir sua própria conta.")
    existing = await _turso_query("SELECT id FROM ac_users WHERE id=?", [user_id])
    if not existing:
        raise HTTPException(404, "Usuário não encontrado.")
    await _turso_exec("DELETE FROM ac_users WHERE id=?", [user_id])


@app.post("/api/admin/users/{user_id}/resend-welcome")
async def admin_resend_welcome(user_id: str, current_user=Depends(_require_admin)):
    rows = await _turso_query("SELECT name, email FROM ac_users WHERE id=?", [user_id])
    if not rows:
        raise HTTPException(404, "Usuário não encontrado.")
    u = rows[0]
    if not _SMTP_HOST:
        raise HTTPException(503, "E-mail não configurado no servidor.")
    await _create_welcome_token_and_send(u["name"], u["email"])
    return {"ok": True}


def _rule_based_analysis(req: AnalyzeRequest, receita: dict, contabil_result: Optional[dict]) -> dict:
    """Gera análise determinística quando a IA não está disponível."""
    d        = receita.get("data", {})
    bureau   = receita.get("status") == "ok"

    # ── Dados Receita Federal ─────────────────────────────────────────────────
    razao       = d.get("razao_social") or req.empresa or "Empresa"
    situacao    = (d.get("descricao_situacao_cadastral") or "").upper()
    status_ativo = "ATIVA" in situacao
    capital_social = float(d.get("capital_social") or 0)
    abertura    = d.get("data_inicio_atividade", "")
    anos_op     = 0
    try:
        anos_op = (date.today() - datetime.strptime(abertura[:10], "%Y-%m-%d").date()).days // 365
    except Exception:
        pass
    simples     = bool(d.get("opcao_pelo_simples"))
    mei         = bool(d.get("opcao_pelo_mei"))
    porte       = d.get("descricao_porte") or ""
    cnae_desc   = d.get("cnae_fiscal_descricao") or req.ramo or "Comércio Exterior"
    uf          = d.get("uf") or ""
    qsa         = d.get("qsa") or []

    # ── Indicadores contábeis ─────────────────────────────────────────────────
    has_contabil   = bool(contabil_result and contabil_result.get("periodo2"))
    p2             = (contabil_result or {}).get("periodo2") or {}
    i2             = p2.get("indicadores") or {}
    t2             = p2.get("totais") or {}
    cr2            = p2.get("credito") or {}
    periodo_label  = (contabil_result or {}).get("periodo2_label") or ""

    def _iv(key):
        iv = i2.get(key) or {}
        v = iv.get("valor_pct") if iv.get("valor_pct") is not None else iv.get("valor")
        return float(v) if v is not None else None

    liq_geral  = _iv("liquidez_geral")
    liq_seca   = _iv("liquidez_seca")
    endiv      = _iv("endividamento_geral")   # % esperado
    mg_liq     = _iv("margem_liquida")         # %
    roe        = _iv("roe")                    # %
    ativo_tot  = t2.get("ativo_total")
    pl         = t2.get("patrimonio_liquido")
    lucro_liq  = t2.get("lucro_liquido")
    rec_liq    = t2.get("receita_liquida")
    cred_calc  = cr2.get("calculado")          # limite calculado pelo sistema

    # ── Limites solicitados ───────────────────────────────────────────────────
    lim_exp    = br_to_float(req.limiteExportador)
    lim_desp   = br_to_float(req.limiteDesp)
    lim_imp    = br_to_float(req.limiteImp)
    exp_total  = lim_exp + lim_desp + lim_imp

    def fmtbr(v):
        if v is None: return "—"
        return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # ── Cálculo de pontuação ──────────────────────────────────────────────────
    score = 50

    # Receita Federal
    if not bureau:
        score -= 5                          # sem dados RF, leve penalidade
    elif not status_ativo:
        score -= 30                         # situação irregular é crítica
    else:
        score += 10                         # empresa ativa

    if anos_op >= 5:
        score += 10
    elif anos_op >= 2:
        score += 5
    elif anos_op > 0:
        score -= 5                          # empresa muito nova

    if capital_social >= 1_000_000:
        score += 5
    elif capital_social >= 100_000:
        score += 2

    if mei:
        score -= 5                          # MEI tem limites estruturais

    # Indicadores contábeis
    if has_contabil:
        if liq_geral is not None:
            if liq_geral >= 1.5:   score += 12
            elif liq_geral >= 1.2: score += 8
            elif liq_geral >= 1.0: score += 4
            elif liq_geral >= 0.8: score -= 8
            else:                  score -= 18

        if endiv is not None:
            if endiv <= 40:   score += 10
            elif endiv <= 60: score += 4
            elif endiv <= 75: score -= 6
            else:             score -= 15

        if mg_liq is not None:
            if mg_liq >= 8:   score += 8
            elif mg_liq >= 3: score += 4
            elif mg_liq >= 0: score += 1
            else:             score -= 12

        if roe is not None:
            if roe >= 15:  score += 5
            elif roe >= 5: score += 2
            elif roe < 0:  score -= 5

        # Proporcionalidade: limite solicitado vs crédito calculado
        if cred_calc and exp_total > 0:
            ratio = exp_total / cred_calc
            if ratio <= 0.5:   score += 5
            elif ratio <= 1.0: score += 2
            elif ratio <= 1.5: score -= 5
            else:              score -= 12

    score = max(0, min(100, score))

    # ── Classificação ────────────────────────────────────────────────────────
    if score >= 90:   classif, recom = "AA", "aprovar"
    elif score >= 80: classif, recom = "A",  "aprovar"
    elif score >= 70: classif, recom = "B",  "aprovar"
    elif score >= 60: classif, recom = "C",  "revisar"
    elif score >= 50: classif, recom = "D",  "revisar"
    else:             classif, recom = "E",  "negar"

    if not status_ativo and bureau:
        recom = "negar"

    # ── Pontos positivos, atenção e críticos ──────────────────────────────────
    pontos_pos  = []
    pontos_at   = []
    alertas     = []

    if status_ativo:
        pontos_pos.append(f"Empresa com situação cadastral ATIVA na Receita Federal")
    else:
        alertas.append(f"Situação cadastral: {situacao} — operação de crédito contraindicada")

    if anos_op >= 5:
        pontos_pos.append(f"Histórico operacional consolidado: {anos_op} anos de atividade")
    elif anos_op >= 2:
        pontos_at.append(f"Empresa com {anos_op} anos de operação — histórico ainda em formação")
    elif anos_op > 0:
        alertas.append(f"Empresa com menos de 2 anos de operação ({anos_op} ano{'s' if anos_op!=1 else ''})")

    if capital_social >= 1_000_000:
        pontos_pos.append(f"Capital social robusto: {fmtbr(capital_social)}")
    elif capital_social >= 100_000:
        pontos_at.append(f"Capital social moderado: {fmtbr(capital_social)}")
    elif capital_social > 0:
        pontos_at.append(f"Capital social reduzido: {fmtbr(capital_social)}")

    if simples:
        pontos_at.append("Optante pelo Simples Nacional — porte limitado")
    if mei:
        alertas.append("Empresa enquadrada como MEI — exposição máxima recomendada é muito restrita")

    if has_contabil:
        if liq_geral is not None:
            if liq_geral >= 1.2:
                pontos_pos.append(f"Liquidez geral adequada: {liq_geral:.2f}")
            elif liq_geral >= 1.0:
                pontos_at.append(f"Liquidez geral no limite: {liq_geral:.2f}")
            else:
                alertas.append(f"Liquidez geral insuficiente: {liq_geral:.2f} (< 1)")

        if endiv is not None:
            if endiv <= 60:
                pontos_pos.append(f"Endividamento controlado: {endiv:.1f}%")
            elif endiv <= 75:
                pontos_at.append(f"Endividamento elevado: {endiv:.1f}%")
            else:
                alertas.append(f"Endividamento crítico: {endiv:.1f}% (risco de insolvência)")

        if mg_liq is not None:
            if mg_liq >= 3:
                pontos_pos.append(f"Margem líquida positiva: {mg_liq:.1f}%")
            elif mg_liq >= 0:
                pontos_at.append(f"Margem líquida estreita: {mg_liq:.1f}%")
            else:
                alertas.append(f"Margem líquida negativa: {mg_liq:.1f}% — empresa com prejuízo")

        if lucro_liq is not None and lucro_liq < 0:
            alertas.append(f"Resultado líquido negativo no período: {fmtbr(lucro_liq)}")

        if cred_calc and exp_total > 0:
            ratio = exp_total / cred_calc
            if ratio > 1.5:
                alertas.append(f"Limite solicitado ({fmtbr(exp_total)}) excede em {ratio:.1f}x o crédito calculado ({fmtbr(cred_calc)})")
            elif ratio > 1.0:
                pontos_at.append(f"Limite solicitado levemente acima do crédito calculado ({fmtbr(cred_calc)})")
            else:
                pontos_pos.append(f"Limite solicitado proporcional ao crédito calculado ({fmtbr(cred_calc)})")
    else:
        pontos_at.append("Demonstrativos financeiros não analisados — análise baseada apenas em dados cadastrais")

    # ── Limites recomendados ──────────────────────────────────────────────────
    if cred_calc:
        base = cred_calc * (score / 100)
        prop_exp  = fmtbr(min(lim_exp,  base * 0.70)) if lim_exp  else "—"
        prop_desp = fmtbr(min(lim_desp, base * 0.20)) if lim_desp else "—"
        prop_imp  = fmtbr(min(lim_imp,  base * 0.10)) if lim_imp  else "—"
        exp_rec   = fmtbr(base)
    else:
        prop_exp  = req.limiteExportador or "—"
        prop_desp = req.limiteDesp or "—"
        prop_imp  = req.limiteImp or "—"
        exp_rec   = fmtbr(exp_total) if exp_total else "—"

    # Prazo
    prazo_raw = req.prazoPagtoVendemmia or ""
    try:
        prazo_rec = int(re.search(r"\d+", prazo_raw).group())
    except Exception:
        prazo_rec = 30

    # ── Textos ───────────────────────────────────────────────────────────────
    tempo_op = calc_tempo_mercado(req.fundacao or abertura)

    if has_contabil and periodo_label:
        ctx_contabil = f" Os demonstrativos financeiros do período {periodo_label} foram analisados e refletem-se na pontuação."
    else:
        ctx_contabil = " Não foram apresentados demonstrativos financeiros; a análise baseia-se exclusivamente em dados cadastrais."

    resumo = (
        f"{razao} é uma empresa do segmento {cnae_desc}, com {tempo_op} de atividade e situação "
        f"cadastral {situacao or 'não verificada'} na Receita Federal.{ctx_contabil} "
        f"Com base nos indicadores disponíveis, a empresa recebeu classificação {classif} "
        f"(score {score}/100), com recomendação de {recom.upper()} para a operação de crédito solicitada."
    )

    analise_cad = (
        f"CNPJ consultado em tempo real via BrasilAPI. "
        f"Situação cadastral: {situacao or '—'}. "
        f"Porte: {porte or '—'}. "
        f"Capital social declarado: {fmtbr(capital_social)}. "
        f"Natureza jurídica: {d.get('descricao_natureza_juridica') or '—'}. "
        f"UF sede: {uf}. "
        f"{'Optante pelo Simples Nacional. ' if simples else ''}"
        f"{'Enquadrada como MEI. ' if mei else ''}"
    )

    analise_soc = ""
    if qsa:
        socs = [f"{s.get('nome_socio','?')} ({s.get('percentual_capital_social','?')}%)" for s in qsa[:5]]
        analise_soc = f"Quadro societário identificado: {', '.join(socs)}."
    else:
        analise_soc = "Quadro societário não disponível na consulta pública."

    analise_prop = (
        f"Exposição total solicitada: {fmtbr(exp_total)}. "
        + (f"Crédito máximo calculado pelos demonstrativos: {fmtbr(cred_calc)}. " if cred_calc else "")
        + (f"Proporção: {exp_total/cred_calc:.1f}x o limite calculado. " if cred_calc and exp_total else "")
        + f"Score de {score}/100 aplicado ao dimensionamento dos limites recomendados."
    )

    analise_op = (
        f"Produto/modalidade: {req.produto or '—'}. Modal: {req.modal or '—'}. "
        f"Origens: {req.origens or '—'}. Incoterms: {req.incoterms or '—'}. "
        f"Tipo de operação: {req.tipoOp or '—'}. Prazo pagamento Vendemmia: {prazo_raw or '—'}."
    )

    fundament = (
        f"Análise determinística gerada a partir de dados objetivos — IA temporariamente indisponível. "
        f"Score {score}/100 composto por: dados cadastrais Receita Federal ({'+10' if status_ativo else '-30'} status; "
        f"{'+10' if anos_op>=5 else '+5' if anos_op>=2 else '-5' if anos_op>0 else '0'} tempo de operação)"
        + (f"; indicadores contábeis (liquidez, endividamento, margem)" if has_contabil else "")
        + f". Classificação {classif} segue escala AA→E conforme política interna Vendemmia."
    )

    return {
        "score": score,
        "classificacao": classif,
        "recomendacao": recom,
        "resumo_executivo": resumo,
        "pontos_positivos": pontos_pos,
        "pontos_atencao": pontos_at,
        "alertas_criticos": alertas,
        "fundamentacao": fundament,
        "analise_cadastral": analise_cad,
        "analise_societaria": analise_soc,
        "analise_proporcionalidade": analise_prop,
        "analise_operacional": analise_op,
        "limite_recomendado_exportador": prop_exp,
        "limite_recomendado_desp": prop_desp,
        "limite_recomendado_imp": prop_imp,
        "exposicao_total_recomendada": exp_rec,
        "prazo_recomendado": prazo_rec,
    }


# ── Endpoints de análise ──────────────────────────────────────────────────────


@app.get("/api/receita/{cnpj}")
@limiter.limit("30/minute")
async def get_receita(request: Request, cnpj: str, current_user=Depends(_get_current_user)):
    """Consulta isolada à Receita Federal via BrasilAPI — independente da IA."""
    return await fetch_receita(cnpj)


@app.post("/api/analyze")
@limiter.limit("30/minute")
async def analyze(request: Request, req: AnalyzeRequest, current_user=Depends(_get_current_user)):
    # 1. Consulta Receita Federal — sempre executada, nunca bloqueia a resposta
    receita = await fetch_receita(req.cnpj)

    # 1b. Carrega indicadores contábeis se sol_id informado
    contabil_result: Optional[dict] = None
    if req.sol_id and _SOL_ID_RE.match(req.sol_id) and _turso_ok():
        try:
            sol_rows = await _turso_query("SELECT data FROM ac_solicitacoes WHERE id=?", [req.sol_id])
            if sol_rows:
                sol_data = json.loads(sol_rows[0]["data"] or "{}")
                contabil_result = sol_data.get("contabil_result")
        except Exception:
            pass

    # 2. Tenta análise com Gemini — falha de forma isolada
    key = _load_key()
    analysis: Optional[Dict[str, Any]] = None
    ai_error: Optional[str] = None

    if not key:
        ai_error = "GEMINI_API_KEY não configurada no servidor."
    else:
        prompt = build_prompt(req, receita, contabil_result)
        try:
            raw = await asyncio.to_thread(_gemini_generate, key, prompt)

            # 3. Extrai JSON com três estratégias em cascata
            analysis = _extract_json(raw)
            if analysis is None:
                analysis = {
                    "score": 0,
                    "classificacao": "—",
                    "recomendacao": "revisar",
                    "resumo_executivo": raw,
                    "pontos_positivos": [],
                    "pontos_atencao": [],
                    "alertas_criticos": ["Erro ao processar resposta da IA — revise manualmente"],
                    "fundamentacao": raw,
                    "analise_cadastral": "",
                    "analise_societaria": "",
                    "analise_proporcionalidade": "",
                    "analise_operacional": "",
                    "limite_recomendado_exportador": "—",
                    "limite_recomendado_desp": "—",
                    "limite_recomendado_imp": "—",
                    "exposicao_total_recomendada": "—",
                    "prazo_recomendado": 30,
                }

        except Exception as exc:
            ai_error = f"[{type(exc).__name__}] {exc}"

    # Fallback determinístico quando a IA não retornou resultado
    ai_fallback = False
    if analysis is None:
        try:
            analysis = _rule_based_analysis(req, receita, contabil_result)
            ai_fallback = True
        except Exception:
            pass

    return {
        "cnpj_data": receita,
        "analysis": analysis,
        "ai_error": ai_error,
        "ai_fallback": ai_fallback,
        "bureau_fonte": "BrasilAPI / Receita Federal",
        "modelo_ia": _load_gemini_model(),
    }


# ── CRUD Solicitações ────────────────────────────────────────────────────────

@app.get("/api/solicitacoes")
async def sol_list(current_user=Depends(_get_current_user)):
    rows = await _turso_query(
        "SELECT id, status, created_at, updated_at, created_by, data FROM ac_solicitacoes ORDER BY created_at DESC"
    )
    items = []
    for row in rows:
        try:
            d = json.loads(row["data"] or "{}")
        except Exception:
            d = {}
        d["id"]        = row["id"]
        d["status"]    = row["status"]
        d["createdAt"] = row["created_at"]
        d["updatedAt"] = row["updated_at"]
        try:
            d["created_by"] = json.loads(row["created_by"] or "null")
        except Exception:
            pass
        if _record_visible_to(d, current_user):
            items.append(d)
    return {"items": items}


@app.post("/api/solicitacoes", status_code=201)
async def sol_create(request: Request, current_user=Depends(_get_current_user)):
    body       = await request.json()
    sol_id     = body.get("id") or str(uuid.uuid4())
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    status     = body.get("status", "pendente")
    created_at = body.get("createdAt") or datetime.utcnow().isoformat()
    updated_at = body.get("updatedAt") or created_at
    created_by = json.dumps({
        "id": current_user.get("sub"), "email": current_user.get("email"), "name": current_user.get("name")
    })
    body["id"]        = sol_id
    body["createdAt"] = created_at
    body["updatedAt"] = updated_at
    await _turso_exec(
        "INSERT INTO ac_solicitacoes (id, status, created_at, updated_at, created_by, data)"
        " VALUES (?,?,?,?,?,?)"
        " ON CONFLICT (id) DO UPDATE SET status=EXCLUDED.status, created_at=EXCLUDED.created_at,"
        " updated_at=EXCLUDED.updated_at, created_by=EXCLUDED.created_by, data=EXCLUDED.data",
        [sol_id, status, created_at, updated_at, created_by, json.dumps(body, ensure_ascii=False)],
    )

    # Disparo automático de e-mail para os analistas ao criar nova solicitação
    if _NOTIFY_EMAILS:
        empresa     = body.get("empresa") or body.get("nomeEmpresa") or body.get("razaoSocial") or "—"
        cnpj        = body.get("cnpj") or ""
        solicitante = current_user.get("name", "—")
        now_str     = datetime.now().strftime("%d/%m/%Y às %H:%M")
        subject     = f"[Nova Solicitação] {empresa}"
        html        = _email_html("Nova Solicitação de Crédito", "#6366f1", [
            ("Empresa",     empresa),
            ("CNPJ",        cnpj),
            ("Solicitante", solicitante),
            ("Data/hora",   now_str),
        ])
        asyncio.create_task(_send_email(
            subject, html, _NOTIFY_EMAILS,
            from_name=solicitante,
            from_email=current_user.get("email", ""),
        ))

    return {"ok": True, "id": sol_id}


@app.get("/api/solicitacoes/stats")
async def sol_stats(current_user=Depends(_get_current_user)):
    rows   = await _turso_query("SELECT status, COUNT(*) as cnt FROM ac_solicitacoes GROUP BY status")
    counts = {"total": 0, "aprovado": 0, "negado": 0, "em_analise": 0, "pendente": 0, "em_comite": 0}
    for row in rows:
        st  = row["status"] or "pendente"
        cnt = int(row["cnt"] or 0)
        counts["total"] += cnt
        if st in counts:
            counts[st] = cnt
    return counts


@app.get("/api/solicitacoes/{sol_id}")
async def sol_get(sol_id: str, current_user=Depends(_get_current_user)):
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    rows = await _turso_query(
        "SELECT id, status, created_at, updated_at, created_by, data FROM ac_solicitacoes WHERE id=?", [sol_id]
    )
    if not rows:
        raise HTTPException(404, "Solicitação não encontrada")
    row = rows[0]
    try:
        d = json.loads(row["data"] or "{}")
    except Exception:
        d = {}
    d["id"]        = row["id"]
    d["status"]    = row["status"]
    d["createdAt"] = row["created_at"]
    d["updatedAt"] = row["updated_at"]
    if not _record_visible_to(d, current_user):
        raise HTTPException(403, "Acesso negado")
    return d


@app.put("/api/solicitacoes/{sol_id}")
async def sol_update(sol_id: str, request: Request, current_user=Depends(_get_current_user)):
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    body       = await request.json()
    status     = body.get("status", "pendente")
    updated_at = body.get("updatedAt") or datetime.utcnow().isoformat()
    created_at = body.get("createdAt") or updated_at
    existing   = await _turso_query("SELECT created_by, created_at FROM ac_solicitacoes WHERE id=?", [sol_id])
    if existing:
        created_by = existing[0]["created_by"]
        created_at = existing[0]["created_at"] or created_at
    else:
        created_by = json.dumps({
            "id": current_user.get("sub"), "email": current_user.get("email"), "name": current_user.get("name")
        })
    body["id"]        = sol_id
    body["updatedAt"] = updated_at
    await _turso_exec(
        "INSERT INTO ac_solicitacoes (id, status, created_at, updated_at, created_by, data)"
        " VALUES (?,?,?,?,?,?)"
        " ON CONFLICT (id) DO UPDATE SET status=EXCLUDED.status, created_at=EXCLUDED.created_at,"
        " updated_at=EXCLUDED.updated_at, created_by=EXCLUDED.created_by, data=EXCLUDED.data",
        [sol_id, status, created_at, updated_at, created_by, json.dumps(body, ensure_ascii=False)],
    )
    return {"ok": True}


@app.delete("/api/solicitacoes/{sol_id}", status_code=204)
async def sol_delete(sol_id: str, current_user=Depends(_get_current_user)):
    if not _SOL_ID_RE.match(sol_id):
        raise HTTPException(400, "ID inválido")
    if not _user_can_decide(current_user):
        raise HTTPException(403, "Apenas analistas podem excluir solicitações")
    await _turso_exec("DELETE FROM ac_solicitacoes WHERE id=?", [sol_id])
    from fastapi.responses import Response as _R
    return _R(status_code=204)


# ── Admin: Backup e Exportação ────────────────────────────────────────────────

def _build_backup_zip() -> io.BytesIO:
    """Comprime historico/ + docs/ + users.json num ZIP em memória."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(HISTORICO_DIR.glob("*.json")):
            zf.write(f, f"historico/{f.name}")
        for f in DOCS_DIR.rglob("*"):
            if f.is_file():
                zf.write(f, f"docs/{f.relative_to(DOCS_DIR)}")
        if _USERS_FILE.exists():
            zf.write(_USERS_FILE, "users.json")
        # Manifesto com metadados do backup
        manifest = {
            "gerado_em": datetime.now().isoformat(),
            "versao_api": "2.0.0",
            "total_analises": len(list(HISTORICO_DIR.glob("*.json"))),
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    buf.seek(0)
    return buf


@app.get("/api/admin/backup/download")
@limiter.limit("5/hour")
async def admin_backup_download(request: Request, current_user=Depends(_require_admin)):
    """Baixa um ZIP completo com todos os dados (historico, docs, usuários)."""
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    buf = _build_backup_zip()
    # Persiste cópia local na pasta backups/
    local_path = BACKUPS_DIR / f"backup_{ts}.zip"
    local_path.write_bytes(buf.read())
    buf.seek(0)
    # Rotação: mantém os 30 backups mais recentes em disco
    all_backups = sorted(BACKUPS_DIR.glob("backup_*.zip"))
    for old in all_backups[:-30]:
        old.unlink()
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=backup_vendemmia_{ts}.zip"},
    )


@app.get("/api/admin/backup/list")
async def admin_backup_list(current_user=Depends(_require_admin)):
    """Lista os backups armazenados localmente."""
    files = sorted(BACKUPS_DIR.glob("backup_*.zip"), reverse=True)
    return {
        "backups": [
            {
                "nome":       f.name,
                "tamanho_kb": round(f.stat().st_size / 1024, 1),
                "criado_em":  datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            }
            for f in files
        ]
    }


@app.get("/api/admin/export")
async def admin_export_json(current_user=Depends(_require_admin)):
    """Exporta todos os dados em formato normalizado para migração ao banco de dados.

    Retorna três coleções prontas para INSERT em tabelas SQL:
      - analises     → tabela principal (uma linha por análise)
      - decisoes     → decisões do analista (FK: analise_id)
      - documentos   → arquivos enviados (FK: sol_id)
    """
    analises, decisoes, documentos = [], [], []

    for f in sorted(HISTORICO_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime):
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue

        ai  = d.get("analise_ia")  or {}
        rf  = d.get("receita_federal") or {}
        ts  = d.get("timestamps") or {}
        cb  = d.get("created_by") or {}
        dec = d.get("decisao_analista") or {}

        analises.append({
            "id":                    d.get("id"),
            "solicitacao_id":        d.get("solicitacao_id"),
            "empresa":               d.get("empresa"),
            "cnpj":                  d.get("cnpj"),
            "status":                d.get("status_solicitacao"),
            "solicitante_nome":      d.get("solicitante"),
            "solicitante_id":        cb.get("id"),
            "solicitante_email":     cb.get("email"),
            "score_ia":              ai.get("score"),
            "classificacao_ia":      ai.get("classificacao"),
            "recomendacao_ia":       ai.get("recomendacao"),
            "exposicao_recomendada": ai.get("exposicao_total_recomendada"),
            "modelo_ia":             d.get("modelo_ia"),
            "rf_situacao":           (rf.get("data") or {}).get("descricao_situacao_cadastral"),
            "rf_abertura":           (rf.get("data") or {}).get("data_inicio_atividade"),
            "criado_em":             ts.get("solicitacao_criada_at") or d.get("data_solicitacao"),
            "rf_consultada_em":      ts.get("rf_consultada_at"),
            "analise_ia_em":         ts.get("analise_ia_at"),
            "salvo_em":              ts.get("historico_salvo_at"),
        })

        if dec and dec.get("status"):
            decisoes.append({
                "analise_id":      d.get("id"),
                "status":          dec.get("status"),
                "limite_aprovado": dec.get("limiteAprovado"),
                "limite_desp":     dec.get("limiteDesp"),
                "limite_imp":      dec.get("limiteImp"),
                "prazo_aprovado":  dec.get("prazoAprovado"),
                "obs_analista":    dec.get("analistaObs"),
                "parecer_tecnico": dec.get("parecerTecnico"),
                "decisao_analista":dec.get("decisaoAnalista"),
                "decidido_em":     dec.get("decisao_at"),
            })

    for sol_dir in DOCS_DIR.iterdir():
        if not sol_dir.is_dir():
            continue
        for tipo_dir in sol_dir.iterdir():
            if not tipo_dir.is_dir():
                continue
            for arq in tipo_dir.iterdir():
                if arq.is_file():
                    documentos.append({
                        "sol_id":       sol_dir.name,
                        "tipo":         tipo_dir.name,
                        "nome_arquivo": arq.name,
                        "tamanho_bytes":arq.stat().st_size,
                        "path_relativo":str(arq.relative_to(DOCS_DIR)),
                    })

    return {
        "exportado_em": datetime.now().isoformat(),
        "totais": {
            "analises":   len(analises),
            "decisoes":   len(decisoes),
            "documentos": len(documentos),
        },
        "analises":   analises,
        "decisoes":   decisoes,
        "documentos": documentos,
    }


@app.get("/api/health")
async def health():
    """Diagnóstico de conectividade — não requer autenticação."""
    pg_ok = _PG_POOL is not None
    pg_detail: dict = {}
    if pg_ok:
        try:
            async with _PG_POOL.acquire() as conn:
                count = await conn.fetchval("SELECT COUNT(*) FROM ac_solicitacoes")
            pg_detail = {"ac_solicitacoes": count}
        except Exception as exc:
            pg_detail = {"error": str(exc)}

    # Diagnóstico da chave Gemini — mostra prefixo mascarado e validade aparente
    gemini_key = _load_key()
    gemini_ok  = bool(gemini_key) and gemini_key not in ("sua-chave-aqui", "")
    gemini_info: dict = {
        "configured": gemini_ok,
        "length":     len(gemini_key),
        "prefix":     (gemini_key[:6] + "…") if len(gemini_key) >= 6 else "(vazio)",
        "looks_valid": (gemini_key.startswith("AIzaSy") or gemini_key.startswith("AQ.")) if gemini_key else False,
        "model":       _load_gemini_model(),
    }

    return {
        "status":       "ok" if (pg_ok and gemini_ok) else "degraded",
        "pool":         "connected" if pg_ok else None,
        "pg_host":      _PG_HOST or None,
        "pg_db":        _PG_DB   or None,
        "pg_pass_set":  bool(_PG_PASS),
        "gemini":       gemini_info,
        **pg_detail,
    }


# Serve os arquivos HTML/JS/CSS estáticos na raiz
_static_dir = os.path.join(os.path.dirname(__file__), "..")
if os.path.isdir(_static_dir):
    app.mount("/", StaticFiles(directory=_static_dir, html=True), name="static")
